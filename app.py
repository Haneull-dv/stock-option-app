import os
import secrets
import json
import tempfile
from datetime import datetime
from flask import (
    Flask, render_template, request, redirect, url_for,
    jsonify, send_from_directory, send_file, abort, session
)
from werkzeug.utils import secure_filename
import openpyxl

import database as db
from processors.pdf_merger import merge_docs_by_type, get_pdf_page_count
from processors.hwpx_writer import generate_hwpx
from processors.number_korean import number_to_korean
from processors.excel_writer import generate_exercise_excel, generate_registration_excel, generate_issuance_detail_excel
from processors.ocr_reader import extract_rrn_batch
from processors.docx_writer import generate_hwakjakseo, generate_gongmun
from processors.pdf_name_extractor import extract_name_from_pdf, match_name_to_applicants
from processors.step04_generator import generate_step04_documents
from processors.step06_generator import generate_step06_zip

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__)
app.secret_key = secrets.token_hex(32)

# ── Config ─────────────────────────────────────────────────────────────────────
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'data', 'uploads')
OUTPUT_FOLDER = os.path.join(BASE_DIR, 'data', 'outputs')
MAX_CONTENT_LENGTH = 50 * 1024 * 1024  # 50 MB
ALLOWED_EXTENSIONS = {'pdf', 'jpg', 'jpeg', 'png', 'hwp', 'hwpx', 'doc', 'docx'}

TEMPLATES_HWP = os.path.join(BASE_DIR, 'templates_hwp')

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

DOC_TYPE_LABELS = {
    'application': '신청서',
    'id_copy': '신분증사본',
    'account_copy': '계좌사본',
}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def save_uploaded_file(file, round_id, applicant_id, doc_type):
    """Save an uploaded file and return (unique_name, file_path, original_filename).
    확장자는 원본 파일명에서 직접 추출 (secure_filename이 한글 제거하는 문제 방지).
    """
    original = file.filename
    # 원본 파일명에서 확장자 추출
    if '.' in original:
        raw_ext = original.rsplit('.', 1)[1].strip().lower()
        ext = ''.join(c for c in raw_ext if c.isalnum())[:10] or 'bin'
    else:
        ext = 'bin'
    unique_name = f"{doc_type}_{secrets.token_hex(8)}.{ext}"

    dir_path = os.path.join(UPLOAD_FOLDER, str(round_id), str(applicant_id))
    os.makedirs(dir_path, exist_ok=True)
    file_path = os.path.join(dir_path, unique_name)
    file.save(file_path)
    return unique_name, file_path, original


def enrich_applicants(applicants, round_id):
    """Attach doc info to each applicant dict."""
    for ap in applicants:
        docs = db.get_documents(ap['id'])
        doc_map = {d['doc_type']: d for d in docs}
        ap['docs'] = doc_map
        ap['has_application'] = 'application' in doc_map
        ap['has_id_copy'] = 'id_copy' in doc_map
        ap['has_account_copy'] = 'account_copy' in doc_map
    return applicants


# ── Main routes ────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    rounds = db.get_all_rounds()
    for r in rounds:
        stats = db.get_round_stats(r['id'])
        r['total'] = stats['total']
        r['submitted'] = stats['submitted']
        # Simple step progress: check step_outputs
        outputs = db.get_step_outputs(r['id'])
        steps_done = len(set(o['step'] for o in outputs))
        r['steps_done'] = steps_done
    return render_template('index.html', rounds=rounds)


@app.route('/round/new', methods=['GET', 'POST'])
def round_new():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        exercise_date = request.form.get('exercise_date', '').strip()
        notes = request.form.get('notes', '').strip()
        prices = request.form.getlist('prices')
        prices = [p.strip() for p in prices if p.strip().isdigit()]

        if not name:
            return render_template('round_new.html', error='회차명을 입력하세요.')

        round_id = db.create_round(name, exercise_date, notes, prices)
        return redirect(url_for('round_detail', round_id=round_id))

    return render_template('round_new.html')


@app.route('/round/<int:round_id>')
def round_detail(round_id):
    round_obj = db.get_round(round_id)
    if not round_obj:
        abort(404)
    prices = db.get_prices_for_round(round_id)
    stats = db.get_round_stats(round_id)
    outputs = db.get_step_outputs(round_id)
    steps_done = set(o['step'] for o in outputs)
    return render_template(
        'round_detail.html',
        round=round_obj,
        prices=prices,
        stats=stats,
        steps_done=steps_done,
    )


@app.route('/round/<int:round_id>/edit', methods=['POST'])
def round_edit(round_id):
    round_obj = db.get_round(round_id)
    if not round_obj:
        return jsonify(success=False, message='회차를 찾을 수 없습니다.'), 404

    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    exercise_date = (data.get('exercise_date') or '').strip()
    notes = (data.get('notes') or '').strip()
    prices = [p.strip() for p in (data.get('prices') or []) if str(p).strip().isdigit()]

    if not name:
        return jsonify(success=False, message='회차명을 입력하세요.')

    db.update_round(round_id, name, exercise_date, notes, prices)
    return jsonify(success=True)


# ── Step 01 routes ─────────────────────────────────────────────────────────────

@app.route('/round/<int:round_id>/step01')
def step01(round_id):
    round_obj = db.get_round(round_id)
    if not round_obj:
        abort(404)
    prices = db.get_prices_for_round(round_id)
    applicants = db.get_applicants(round_id)
    applicants = enrich_applicants(applicants, round_id)
    outputs = db.get_step_outputs(round_id, 'step01')
    stats = db.get_round_stats(round_id)

    # per-type counts
    type_counts = {}
    for dt in ['application', 'id_copy', 'account_copy']:
        docs = db.get_documents_by_type(round_id, dt)
        type_counts[dt] = len(docs)

    return render_template(
        'step01.html',
        round=round_obj,
        prices=prices,
        applicants=applicants,
        outputs=outputs,
        stats=stats,
        type_counts=type_counts,
        base_url=request.host_url.rstrip('/'),
    )


@app.route('/round/<int:round_id>/applicants/add', methods=['POST'])
def add_applicant(round_id):
    data = request.get_json(silent=True) or request.form
    name = (data.get('name') or '').strip()
    exercise_price = data.get('exercise_price')
    quantity = data.get('quantity')
    broker = (data.get('broker') or '').strip()
    account_number = (data.get('account_number') or '').strip()

    if not name:
        return jsonify(success=False, message='이름을 입력하세요.')

    try:
        exercise_price = int(exercise_price) if exercise_price else None
        quantity = int(quantity) if quantity else None
    except (ValueError, TypeError):
        return jsonify(success=False, message='행사가액/수량은 숫자여야 합니다.')

    applicant_id = db.add_applicant(round_id, name, exercise_price, quantity, broker, account_number)
    ap = db.get_applicant(applicant_id)
    token_link = url_for('employee_submit', token=ap['submit_token'], _external=True)

    return jsonify(
        success=True,
        message='신청자가 추가되었습니다.',
        data={
            'id': applicant_id,
            'name': name,
            'exercise_price': exercise_price,
            'quantity': quantity,
            'broker': broker,
            'account_number': account_number,
            'sort_order': ap['sort_order'],
            'submit_token': ap['submit_token'],
            'token_link': token_link,
            'has_application': False,
            'has_id_copy': False,
            'has_account_copy': False,
        }
    )


@app.route('/round/<int:round_id>/applicants/import-excel', methods=['POST'])
def import_excel(round_id):
    """엑셀 파일에서 신청자 명단 파싱하여 반환 (미리보기용). 실제 저장은 /confirm 으로."""
    if 'file' not in request.files:
        return jsonify(success=False, message='파일이 없습니다.')
    file = request.files['file']
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify(success=False, message='엑셀 파일(.xlsx)만 업로드 가능합니다.')

    # 임시 저장 후 파싱
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    file.save(tmp.name)
    tmp.close()

    try:
        wb = openpyxl.load_workbook(tmp.name, data_only=True)
    except Exception as e:
        os.unlink(tmp.name)
        return jsonify(success=False, message=f'파일을 열 수 없습니다: {str(e)}')
    finally:
        try: os.unlink(tmp.name)
        except: pass

    # 시트 선택: '이름' 헤더가 있는 시트 우선, 없으면 첫 번째 시트
    target_ws = None
    col_map = {}  # field -> col_index (1-based)

    # 헤더 키워드 매핑
    HEADER_KEYS = {
        '이름': 'name',
        '부여일': 'grant_date',
        '행사가': 'exercise_price',
        '주식수': 'total_qty',
        '신청': 'quantity',
        '행사일': 'exercise_date',
        '증권사': 'broker',
        '계좌': 'account_number',
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            # B열(index 1)에 '이름' 텍스트가 있는 행을 헤더로
            row_list = list(row)
            if len(row_list) < 2:
                continue
            # B열 = index 1
            b_val = str(row_list[1]).strip() if row_list[1] is not None else ''
            if b_val == '이름':
                # 이 행이 헤더
                for ci, cell_val in enumerate(row_list):
                    if cell_val is None:
                        continue
                    cell_str = str(cell_val).strip()
                    for keyword, field in HEADER_KEYS.items():
                        if keyword in cell_str and field not in col_map:
                            col_map[field] = ci  # 0-based index
                target_ws = ws
                header_row = row_idx
                break
        if target_ws:
            break

    if not target_ws or 'name' not in col_map:
        return jsonify(success=False, message='이름 헤더를 찾을 수 없습니다. 헤더 행에 "이름" 열이 있는지 확인하세요.')

    # 데이터 파싱
    parsed = []
    for row in target_ws.iter_rows(min_row=header_row + 1, values_only=True):
        row_list = list(row)
        name_val = row_list[col_map['name']] if col_map['name'] < len(row_list) else None
        if not name_val or str(name_val).strip() == '':
            continue  # 이름 없으면 skip
        name = str(name_val).strip()

        def get_col(field, default=None):
            if field not in col_map:
                return default
            idx = col_map[field]
            if idx >= len(row_list):
                return default
            v = row_list[idx]
            return v if v is not None else default

        # 행사가
        try:
            price = int(get_col('exercise_price', 0) or 0)
        except (ValueError, TypeError):
            price = None

        # 신청수량 (quantity)
        try:
            qty = int(get_col('quantity', 0) or 0)
        except (ValueError, TypeError):
            qty = None

        # 부여일
        grant_date = get_col('grant_date')
        if grant_date and hasattr(grant_date, 'strftime'):
            grant_date = grant_date.strftime('%Y-%m-%d')
        elif grant_date:
            grant_date = str(grant_date)[:10]

        # 행사일
        exercise_date = get_col('exercise_date')
        if exercise_date and hasattr(exercise_date, 'strftime'):
            exercise_date = exercise_date.strftime('%Y-%m-%d')
        elif exercise_date:
            exercise_date = str(exercise_date)[:10]

        broker = str(get_col('broker', '') or '').strip()
        account = str(get_col('account_number', '') or '').strip()

        parsed.append({
            'name': name,
            'grant_date': grant_date,
            'exercise_price': price,
            'quantity': qty,
            'exercise_date': exercise_date,
            'broker': broker,
            'account_number': account,
        })

    if not parsed:
        return jsonify(success=False, message='데이터 행을 찾을 수 없습니다.')

    # 회차의 행사일로 필터링
    round_obj = db.get_round(round_id)
    round_exercise_date = (round_obj.get('exercise_date') or '').strip()[:10]  # YYYY-MM-DD

    if round_exercise_date:
        filtered = [r for r in parsed if (r.get('exercise_date') or '')[:10] == round_exercise_date]
        filtered_out = len(parsed) - len(filtered)
    else:
        filtered = parsed
        filtered_out = 0

    if not filtered:
        return jsonify(
            success=False,
            message=f'행사일 {round_exercise_date}에 해당하는 행이 없습니다. '
                    f'(전체 {len(parsed)}행 중 0건 일치) '
                    f'회차 행사일 설정을 확인하세요.'
        )

    return jsonify(success=True, data=filtered, count=len(filtered),
                   filtered_out=filtered_out, exercise_date=round_exercise_date)


@app.route('/round/<int:round_id>/applicants/import-confirm', methods=['POST'])
def import_confirm(round_id):
    """파싱된 신청자 목록을 실제 DB에 저장."""
    data = request.get_json(silent=True) or {}
    applicants_data = data.get('applicants', [])
    mode = data.get('mode', 'append')  # 'append' or 'replace'

    if mode == 'replace':
        db.delete_all_applicants(round_id)

    added = 0
    for ap in applicants_data:
        name = str(ap.get('name', '')).strip()
        if not name:
            continue
        try:
            price = int(ap.get('exercise_price') or 0) or None
            qty   = int(ap.get('quantity') or 0) or None
        except (ValueError, TypeError):
            price, qty = None, None
        broker  = str(ap.get('broker') or '').strip()
        account = str(ap.get('account_number') or '').strip()
        grant_date = str(ap.get('grant_date') or '').strip() or None
        db.add_applicant(round_id, name, price, qty, broker, account, grant_date)
        added += 1

    return jsonify(success=True, message=f'{added}명 추가 완료', added=added)


@app.route('/round/<int:round_id>/applicants/reorder', methods=['POST', 'PATCH'])
def reorder_applicants(round_id):
    data = request.get_json(silent=True) or {}
    id_list = data.get('order', [])
    if not id_list:
        return jsonify(success=False, message='순서 정보가 없습니다.')
    db.reorder_applicants(round_id, [int(i) for i in id_list])
    return jsonify(success=True, message='순서가 업데이트되었습니다.')


@app.route('/round/<int:round_id>/applicants/<int:applicant_id>', methods=['DELETE'])
def delete_applicant(round_id, applicant_id):
    ap = db.get_applicant(applicant_id)
    if not ap or ap['round_id'] != round_id:
        return jsonify(success=False, message='신청자를 찾을 수 없습니다.'), 404
    db.delete_applicant(applicant_id)
    return jsonify(success=True, message='삭제되었습니다.')


@app.route('/round/<int:round_id>/upload/<int:applicant_id>/<doc_type>', methods=['POST'])
def upload_single(round_id, applicant_id, doc_type):
    if doc_type not in DOC_TYPE_LABELS:
        return jsonify(success=False, message='알 수 없는 서류 유형입니다.'), 400

    ap = db.get_applicant(applicant_id)
    if not ap or ap['round_id'] != round_id:
        return jsonify(success=False, message='신청자를 찾을 수 없습니다.'), 404

    if 'file' not in request.files:
        return jsonify(success=False, message='파일이 없습니다.'), 400

    file = request.files['file']
    if not file or not file.filename:
        return jsonify(success=False, message='파일을 선택하세요.'), 400
    if not allowed_file(file.filename):
        return jsonify(success=False, message='허용되지 않는 파일 형식입니다.'), 400

    unique_name, file_path, original = save_uploaded_file(file, round_id, applicant_id, doc_type)
    db.add_document(applicant_id, doc_type, unique_name, original, file_path)

    return jsonify(
        success=True,
        message='업로드 완료',
        data={
            'applicant_id': applicant_id,
            'doc_type': doc_type,
            'original_filename': original,
        }
    )


@app.route('/round/<int:round_id>/applicant/<int:applicant_id>/documents', methods=['GET'])
def get_applicant_documents(round_id, applicant_id):
    """특정 신청자의 모든 서류 목록 조회"""
    ap = db.get_applicant(applicant_id)
    if not ap or ap['round_id'] != round_id:
        return jsonify(success=False, message='신청자를 찾을 수 없습니다.'), 404

    docs = db.get_documents(applicant_id)
    return jsonify(success=True, documents=docs)


@app.route('/round/<int:round_id>/document/<int:doc_id>', methods=['DELETE'])
def delete_document(round_id, doc_id):
    """서류 삭제 (파일 + DB)"""
    conn = db.get_db()
    row = conn.execute(
        """SELECT d.*, a.round_id
           FROM documents d
           JOIN applicants a ON d.applicant_id = a.id
           WHERE d.id=?""",
        (doc_id,)
    ).fetchone()
    conn.close()

    if not row:
        return jsonify(success=False, message='서류를 찾을 수 없습니다.'), 404

    if row['round_id'] != round_id:
        return jsonify(success=False, message='잘못된 요청입니다.'), 400

    # 실제 파일 삭제
    file_path = row['file_path']
    if file_path and os.path.exists(file_path):
        try:
            os.remove(file_path)
        except Exception as e:
            print(f"파일 삭제 실패: {file_path}, {e}")

    # DB에서 삭제
    db.delete_document(doc_id)

    return jsonify(success=True, message='서류가 삭제되었습니다.')


@app.route('/round/<int:round_id>/extract_name_from_pdf', methods=['POST'])
def extract_pdf_name(round_id):
    """PDF에서 신청자 이름 추출 및 매칭"""
    round_obj = db.get_round(round_id)
    if not round_obj:
        return jsonify(success=False, message='회차를 찾을 수 없습니다.'), 404

    if 'file' not in request.files:
        return jsonify(success=False, message='파일이 없습니다.'), 400

    file = request.files['file']
    if not file or not file.filename:
        return jsonify(success=False, message='파일을 선택하세요.'), 400

    # 임시 파일로 저장
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
        file.save(tmp.name)
        tmp_path = tmp.name

    try:
        # PDF에서 이름 추출
        extracted_name = extract_name_from_pdf(tmp_path)

        if not extracted_name:
            return jsonify(success=False, message='PDF에서 이름을 찾을 수 없습니다.')

        # 신청자 명단 조회
        applicants = db.get_applicants(round_id)
        applicant_names = {str(ap['id']): ap['name'] for ap in applicants}

        # 이름 매칭
        matched_id, matched_name = match_name_to_applicants(extracted_name, applicant_names)

        if not matched_id:
            return jsonify(
                success=False,
                message=f'"{extracted_name}"과 일치하는 신청자를 찾을 수 없습니다.',
                extracted_name=extracted_name
            )

        return jsonify(
            success=True,
            extracted_name=extracted_name,
            matched_id=int(matched_id),
            matched_name=matched_name
        )

    except Exception as e:
        return jsonify(success=False, message=f'처리 오류: {str(e)}')

    finally:
        # 임시 파일 삭제
        try:
            os.unlink(tmp_path)
        except:
            pass


@app.route('/round/<int:round_id>/upload/bulk', methods=['POST'])
def upload_bulk(round_id):
    """Bulk upload: files + optional doc_type or auto-detect from filename."""
    round_obj = db.get_round(round_id)
    if not round_obj:
        abort(404)

    applicants = db.get_applicants(round_id)
    files = request.files.getlist('files')
    doc_type_override = request.form.get('doc_type')  # may be empty

    results = []
    for file in files:
        if not file or not file.filename:
            continue
        if not allowed_file(file.filename):
            results.append({'filename': file.filename, 'success': False, 'message': '허용되지 않는 형식'})
            continue

        # Determine doc_type
        fname_lower = file.filename.lower()
        if doc_type_override and doc_type_override in DOC_TYPE_LABELS:
            doc_type = doc_type_override
        elif '신청서' in file.filename or 'application' in fname_lower:
            doc_type = 'application'
        elif '신분증' in file.filename or 'id' in fname_lower:
            doc_type = 'id_copy'
        elif '계좌' in file.filename or 'account' in fname_lower:
            doc_type = 'account_copy'
        else:
            doc_type = doc_type_override or 'application'

        # Try to match applicant name in filename
        matched_ap = None
        for ap in applicants:
            if ap['name'] in file.filename:
                matched_ap = ap
                break

        if not matched_ap:
            # No match — store as unassigned (use a placeholder applicant_id=0 dir)
            dir_path = os.path.join(UPLOAD_FOLDER, str(round_id), 'unassigned')
            os.makedirs(dir_path, exist_ok=True)
            safe = secure_filename(file.filename)
            unique_name = f"{doc_type}_{secrets.token_hex(6)}_{safe}"
            file_path = os.path.join(dir_path, unique_name)
            file.save(file_path)
            results.append({
                'filename': file.filename,
                'success': True,
                'matched': False,
                'message': '미매칭 (수동 배정 필요)',
                'stored_path': file_path,
                'doc_type': doc_type,
            })
        else:
            unique_name, file_path, original = save_uploaded_file(
                file, round_id, matched_ap['id'], doc_type
            )
            db.add_document(matched_ap['id'], doc_type, unique_name, original, file_path)
            results.append({
                'filename': file.filename,
                'success': True,
                'matched': True,
                'applicant_id': matched_ap['id'],
                'applicant_name': matched_ap['name'],
                'doc_type': doc_type,
                'message': f"{matched_ap['name']} - {DOC_TYPE_LABELS[doc_type]} 배정 완료",
            })

    return jsonify(success=True, data=results)


@app.route('/round/<int:round_id>/applicants/status')
def applicants_status(round_id):
    status = db.get_submission_status(round_id)
    return jsonify(success=True, data=status)


@app.route('/round/<int:round_id>/step01/merge', methods=['POST'])
def merge_step01(round_id):
    round_obj = db.get_round(round_id)
    if not round_obj:
        return jsonify(success=False, message='회차를 찾을 수 없습니다.'), 404

    applicants = db.get_applicants(round_id)
    output_dir = os.path.join(OUTPUT_FOLDER, str(round_id), 'step01')
    os.makedirs(output_dir, exist_ok=True)

    display_names = {
        'application': '신청서_합본.pdf',
        'id_copy':     '신분증사본_합본.pdf',
        'account_copy':'계좌사본_합본.pdf',
    }

    results = {}
    for doc_type in ['application', 'id_copy', 'account_copy']:
        try:
            out_path = merge_docs_by_type(
                round_id, doc_type, applicants, UPLOAD_FOLDER, output_dir
            )
            label = DOC_TYPE_LABELS[doc_type]
            filename = os.path.basename(out_path)   # ASCII 파일명
            size = os.path.getsize(out_path) if os.path.exists(out_path) else 0
            pages = get_pdf_page_count(out_path)

            db.save_step_output(round_id, 'step01', filename, out_path)

            results[doc_type] = {
                'success': True,
                'label': label,
                'filename': display_names.get(doc_type, filename),  # UI 표시용 한글명
                'size': size,
                'pages': pages,
                'download_url': url_for(
                    'download_output',
                    round_id=round_id,
                    filename=filename,   # URL에는 ASCII명
                    _external=False
                ),
            }
        except Exception as e:
            results[doc_type] = {
                'success': False,
                'label': DOC_TYPE_LABELS[doc_type],
                'message': str(e),
            }

    return jsonify(success=True, data=results)


@app.route('/round/<int:round_id>/output/<int:output_id>', methods=['DELETE'])
def delete_output(round_id, output_id):
    """결과물 삭제 (파일 + DB)"""
    conn = db.get_db()
    row = conn.execute(
        "SELECT * FROM step_outputs WHERE id=? AND round_id=?",
        (output_id, round_id)
    ).fetchone()
    conn.close()

    if not row:
        return jsonify(success=False, message='결과물을 찾을 수 없습니다.'), 404

    # 실제 파일 삭제
    file_path = row['output_path']
    if file_path and os.path.exists(file_path):
        try:
            os.remove(file_path)
        except Exception as e:
            print(f"파일 삭제 실패: {file_path}, {e}")

    # DB에서 삭제
    db.delete_step_output(output_id)

    return jsonify(success=True, message='결과물이 삭제되었습니다.')


@app.route('/round/<int:round_id>/download/<path:filename>')
def download_output(round_id, filename):
    output_dir = os.path.join(OUTPUT_FOLDER, str(round_id), 'step01')
    safe = os.path.basename(filename)
    full_path = os.path.join(output_dir, safe)
    if not os.path.isfile(full_path):
        abort(404)
    # 한글 표시명 매핑 (실제 파일은 ASCII로 저장)
    display_names = {
        'application_merged.pdf': '신청서_합본.pdf',
        'id_copy_merged.pdf':     '신분증사본_합본.pdf',
        'account_copy_merged.pdf':'계좌사본_합본.pdf',
    }
    display = display_names.get(safe, safe)
    return send_file(full_path, mimetype='application/pdf',
                     as_attachment=True, download_name=display)


# ── Step 03 routes ─────────────────────────────────────────────────────────────

@app.route('/round/<int:round_id>/step03')
def step03(round_id):
    round_obj = db.get_round(round_id)
    if not round_obj:
        abort(404)
    applicants = db.get_applicants(round_id)
    prices = db.get_prices_for_round(round_id)

    # 가격별 집계
    price_summary = {}
    for ap in applicants:
        p = ap.get('exercise_price') or 0
        q = ap.get('quantity') or 0
        if p not in price_summary:
            price_summary[p] = {'qty': 0, 'amount': 0, 'count': 0}
        price_summary[p]['qty']    += q
        price_summary[p]['amount'] += p * q
        price_summary[p]['count']  += 1

    total_qty    = sum(v['qty']    for v in price_summary.values())
    total_amount = sum(v['amount'] for v in price_summary.values())

    outputs = db.get_step_outputs(round_id, 'step03')
    config = db.get_step03_config(round_id)

    return render_template(
        'step03.html',
        round=round_obj,
        prices=sorted(price_summary.keys()),
        price_summary=price_summary,
        total_qty=total_qty,
        total_amount=total_amount,
        outputs=outputs,
        config=config,
        base_url=request.host_url.rstrip('/'),
    )


@app.route('/round/<int:round_id>/step03/config', methods=['POST'])
def step03_config(round_id):
    """Step03 설정 저장."""
    data = request.get_json(silent=True) or {}
    db.save_step03_config(
        round_id,
        payment_date=data.get('payment_date'),
        total_capital=data.get('total_capital')
    )
    return jsonify(success=True)


@app.route('/round/<int:round_id>/step03/generate', methods=['POST'])
def step03_generate(round_id):
    round_obj = db.get_round(round_id)
    if not round_obj:
        return jsonify(success=False, message='회차를 찾을 수 없습니다.'), 404

    data = request.get_json(silent=True) or {}
    payment_date = data.get('payment_date', '').strip()   # e.g. "2026-02-23"
    total_capital = data.get('total_capital', '').strip()  # 자본총액 (수납의뢰서용)
    doc_types = data.get('doc_types', ['excel', 'sunabuiuiseo', 'yeongsujeung', 'bogwan'])

    applicants = db.get_applicants(round_id)

    # 가격별 집계
    price_summary = {}
    for ap in applicants:
        p = ap.get('exercise_price') or 0
        q = ap.get('quantity') or 0
        if p not in price_summary:
            price_summary[p] = {'qty': 0, 'amount': 0}
        price_summary[p]['qty']    += q
        price_summary[p]['amount'] += p * q

    total_qty    = sum(v['qty']    for v in price_summary.values())
    total_amount = sum(v['amount'] for v in price_summary.values())

    # 날짜 파싱
    try:
        dt = datetime.strptime(payment_date, '%Y-%m-%d')
        # 수납의뢰서용 (단일 공백)
        date_kr = f"{dt.year}년 {dt.month:02d}월 {dt.day:02d}일"
        # 영수증용
        date_kr3 = f"{dt.year}년  {dt.month:02d}월  {dt.day:02d}일"
        # 보관증명서용 (XML 태그로 분리된 날짜)
        year_only = str(dt.year)
        date_part = f" 년   {dt.month:02d} 월   {dt.day:02d} 일"
    except ValueError:
        date_kr = date_kr3 = payment_date
        year_only = '2026'
        date_part = ' 년   02 월   23 일'

    output_dir = os.path.join(OUTPUT_FOLDER, str(round_id), 'step03')
    os.makedirs(output_dir, exist_ok=True)

    results = []

    # 1. 행사내역 엑셀
    if 'excel' in doc_types:
        try:
            fname = '주식매수선택권 행사내역.xlsx'
            excel_path = os.path.join(output_dir, fname)
            generate_exercise_excel(
                round_obj['name'],
                round_obj.get('exercise_date', ''),
                applicants,
                excel_path
            )
            db.save_step_output(round_id, 'step03', fname, excel_path)
            results.append({'name': fname, 'filename': fname, 'success': True})
        except Exception as e:
            results.append({'name': '주식납입금 행사내역.xlsx', 'success': False, 'message': str(e)})

    # 2. 수납의뢰서
    if 'sunabuiuiseo' in doc_types:
        try:
            fname = '주식납입금 수납의뢰서.hwpx'
            tpl = os.path.join(TEMPLATES_HWP, '주식납입금 수납의뢰서_260223_스톡옵션.hwpx')
            out = os.path.join(output_dir, fname)
            reps = {
                # 날짜 2개 치환 (한 덩어리 + 분리된 것)
                '2026년 02월 23일': date_kr,
                '2026 ': f'{year_only} ',  # 공백 포함 (두 번째 날짜만 치환)
                '년   02 월   23 일': date_part,
                '110,123': f'{total_qty:,}',
                # 가격 목록은 치환 안 함 (길이 변경 시 HWPX 파일 손상)
            }
            if total_capital:
                reps['5,384,693,000'] = f'{int(total_capital):,}'
            generate_hwpx(tpl, out, reps)
            db.save_step_output(round_id, 'step03', fname, out)
            results.append({'name': fname, 'filename': fname, 'success': True})
        except Exception as e:
            results.append({'name': '주식납입금 수납의뢰서.hwpx', 'success': False, 'message': str(e)})

    # 3. 영수증
    if 'yeongsujeung' in doc_types:
        try:
            fname = '주식납입금 영수증.hwpx'
            tpl = os.path.join(TEMPLATES_HWP, '주식납입금 영수증_260223_스톡옵션.hwpx')
            out = os.path.join(output_dir, fname)
            amount_korean = number_to_korean(total_amount)
            reps = {
                '사천팔백팔십일만천오백': amount_korean,
                '375,577,990': f'{total_amount:,}',
                '2026년  02월  23일': date_kr3,
            }
            generate_hwpx(tpl, out, reps)
            db.save_step_output(round_id, 'step03', fname, out)
            results.append({'name': fname, 'filename': fname, 'success': True})
        except Exception as e:
            results.append({'name': '주식납입금 영수증.hwpx', 'success': False, 'message': str(e)})

    # 4. 보관증명서 (가격별)
    if 'bogwan' in doc_types:
        tpl_map = {
            1250: '주식납입금_보관증명서_발급의뢰서_260223_스톡옵션_1250.hwpx',
            2000: '주식납입금_보관증명서_발급의뢰서_260223_스톡옵션_2000.hwpx',
            4130: '주식납입금_보관증명서_발급의뢰서_260223_스톡옵션_4130.hwpx',
        }
        # 기존 보관증명서 금액 (원본 파일의 텍스트)
        bogwan_orig = {
            1250: {'price': '1,250', 'amount': '13,125,000'},
            2000: {'price': '2,000', 'amount': '46,000,000'},
            4130: {'price': '4,130', 'amount': '316,452,990'},
        }

        for price, info in sorted(price_summary.items()):
            amount = info['amount']
            fname  = f'주식납입금 보관증명서 발급의뢰서_{price:,}원.hwpx'
            out    = os.path.join(output_dir, fname)

            # 템플릿 선택: 동일 가격 있으면 그거, 없으면 가장 가까운 것
            if price in tpl_map:
                tpl = os.path.join(TEMPLATES_HWP, tpl_map[price])
                orig = bogwan_orig[price]
            else:
                # 없는 가격 → 4130 템플릿 기반으로 생성
                closest = min(tpl_map.keys(), key=lambda x: abs(x - price))
                tpl = os.path.join(TEMPLATES_HWP, tpl_map[closest])
                orig = bogwan_orig[closest]

            try:
                reps = {
                    orig['price']: f'{price:,}',
                    orig['amount']: f'{amount:,}',
                    # 보관증명서 날짜는 XML 태그로 분리되어 있어서 두 부분으로 나눠 치환
                    '2026': year_only,
                    ' 년   02 월   23 일': date_part,
                }
                generate_hwpx(tpl, out, reps)
                db.save_step_output(round_id, 'step03', fname, out)
                results.append({
                    'name': fname,
                    'filename': fname,
                    'success': True
                })
            except Exception as e:
                results.append({'name': f'보관증명서_{price:,}원.hwpx', 'success': False, 'message': str(e)})

    # 5. 행사청구서 PDF 합본 (Step01에서 생성된 것 복사)
    if 'haengsa_cheonggu' in doc_types:
        try:
            step01_dir = os.path.join(OUTPUT_FOLDER, str(round_id), 'step01')
            src = os.path.join(step01_dir, 'application_merged.pdf')
            fname = '주식매수선택권 행사청구서.pdf'
            dest = os.path.join(output_dir, fname)

            if os.path.exists(src):
                import shutil
                shutil.copy2(src, dest)
                db.save_step_output(round_id, 'step03', fname, dest)
                results.append({'name': fname, 'filename': fname, 'success': True})
            else:
                results.append({'name': fname, 'success': False, 'message': 'Step01에서 생성된 행사청구서를 찾을 수 없습니다.'})
        except Exception as e:
            results.append({'name': '행사청구서', 'success': False, 'message': str(e)})

    # 6. 정관 (공용 파일 복사)
    if 'jeonggwan' in doc_types:
        try:
            templates_common = os.path.join(BASE_DIR, 'templates_common')
            # 가장 최신 정관 파일 찾기
            jeonggwan_files = [f for f in os.listdir(templates_common) if f.startswith('정관') and f.endswith('.pdf')]
            if jeonggwan_files:
                src = os.path.join(templates_common, jeonggwan_files[0])
                fname = jeonggwan_files[0]
                dest = os.path.join(output_dir, fname)

                import shutil
                shutil.copy2(src, dest)
                db.save_step_output(round_id, 'step03', fname, dest)
                results.append({'name': fname, 'filename': fname, 'success': True})
            else:
                results.append({'name': '정관', 'success': False, 'message': 'templates_common에서 정관 파일을 찾을 수 없습니다.'})
        except Exception as e:
            results.append({'name': '정관', 'success': False, 'message': str(e)})

    # 7. 주주총회의사록 및 조정산식 (발행가액별 + 부여일 필터링 → 폴더 + ZIP)
    if 'jusimchong' in doc_types:
        try:
            import shutil
            import zipfile

            # 발행가액별 신청자 그룹화 및 부여일 수집
            price_grant_dates = {}
            for ap in applicants:
                price = ap.get('exercise_price')
                grant_date = ap.get('grant_date')
                if price and grant_date:
                    if price not in price_grant_dates:
                        price_grant_dates[price] = set()
                    price_grant_dates[price].add(grant_date)

            source_base = os.path.join(BASE_DIR, '주주총회의사록 및 조정산식')
            dest_base = os.path.join(output_dir, '주주총회의사록 및 조정산식')

            # 폴더에 파일 복사
            copied_count = 0
            for price, grant_dates in sorted(price_grant_dates.items()):
                # 발행가액별 폴더 생성
                price_folder = os.path.join(dest_base, f'{price}원')
                os.makedirs(price_folder, exist_ok=True)

                # 원본 폴더 확인
                src_price_folder = os.path.join(source_base, str(price))
                if not os.path.exists(src_price_folder):
                    continue

                # 부여일 매칭 파일 복사
                for grant_date in grant_dates:
                    # grant_date: "YYYY-MM-DD" → "YYYYMMDD"
                    try:
                        date_prefix = grant_date.replace('-', '')  # "20211101"
                    except:
                        continue

                    # 해당 부여일로 시작하는 모든 PDF 복사
                    for filename in os.listdir(src_price_folder):
                        if filename.startswith(date_prefix) and filename.endswith('.pdf'):
                            src_file = os.path.join(src_price_folder, filename)
                            dest_file = os.path.join(price_folder, filename)
                            shutil.copy2(src_file, dest_file)
                            copied_count += 1

            if copied_count > 0:
                # 개별 다운로드용 ZIP 파일 생성
                zip_fname = '주주총회의사록 및 조정산식.zip'
                zip_path = os.path.join(output_dir, zip_fname)

                with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                    for root, dirs, files in os.walk(dest_base):
                        for file in files:
                            file_path = os.path.join(root, file)
                            # ZIP 내부 경로: 주주총회의사록 및 조정산식/4130원/파일명.pdf
                            arc_path = os.path.join('주주총회의사록 및 조정산식',
                                                   os.path.relpath(file_path, dest_base))
                            zf.write(file_path, arc_path)

                db.save_step_output(round_id, 'step03', zip_fname, zip_path)
                results.append({
                    'name': f'주주총회의사록 및 조정산식 ({copied_count}개 파일)',
                    'filename': zip_fname,
                    'success': True
                })
            else:
                results.append({
                    'name': '주주총회의사록 및 조정산식',
                    'success': False,
                    'message': '매칭되는 파일이 없습니다.'
                })
        except Exception as e:
            results.append({'name': '주주총회의사록 및 조정산식', 'success': False, 'message': str(e)})

    return jsonify(success=True, data=results)


@app.route('/round/<int:round_id>/step03/download/<path:filename>')
def download_step03(round_id, filename):
    output_dir = os.path.join(OUTPUT_FOLDER, str(round_id), 'step03')
    safe = os.path.basename(filename)
    full_path = os.path.join(output_dir, safe)
    if not os.path.isfile(full_path):
        abort(404)

    display_map = {
        'exercise_detail.xlsx': '행사내역.xlsx',
        'sunabuiuiseo.hwpx':    '수납의뢰서.hwpx',
        'yeongsujeung.hwpx':    '영수증.hwpx',
    }
    display = display_map.get(safe, safe)
    if safe.startswith('bogwan_'):
        price_str = safe.replace('bogwan_', '').replace('.hwpx', '')
        display = f'보관증명서_{price_str}원.hwpx'

    ext = os.path.splitext(safe)[1].lower()
    mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' if ext == '.xlsx' \
           else 'application/octet-stream'
    return send_file(full_path, mimetype=mime, as_attachment=True, download_name=display)


@app.route('/round/<int:round_id>/step03/download_all_zip')
def download_step03_all_zip(round_id):
    """Step03 전체 서류를 ZIP으로 묶어서 다운로드."""
    import zipfile
    import io
    from datetime import datetime

    round_obj = db.get_round(round_id)
    if not round_obj:
        abort(404)

    output_dir = os.path.join(OUTPUT_FOLDER, str(round_id), 'step03')

    # ZIP 파일을 메모리에 생성
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        # output_dir의 모든 파일 및 폴더를 ZIP에 추가
        if os.path.exists(output_dir):
            for filename in os.listdir(output_dir):
                file_path = os.path.join(output_dir, filename)

                if os.path.isfile(file_path):
                    # 주주총회의사록.zip은 건너뛰기 (폴더로 포함됨)
                    if filename == '주주총회의사록 및 조정산식.zip':
                        continue

                    # 파일인 경우
                    display_map = {
                        'exercise_detail.xlsx': '행사내역.xlsx',
                        'sunabuiuiseo.hwpx': '수납의뢰서.hwpx',
                        'yeongsujeung.hwpx': '영수증.hwpx',
                    }
                    display_name = display_map.get(filename, filename)

                    # bogwan 파일은 가격 포함
                    if filename.startswith('bogwan_'):
                        price_str = filename.replace('bogwan_', '').replace('.hwpx', '')
                        display_name = f'보관증명서_{price_str}원.hwpx'

                    zf.write(file_path, display_name)

                elif os.path.isdir(file_path):
                    # 폴더인 경우 (주주총회의사록 등) - 재귀적으로 추가
                    for root, dirs, files in os.walk(file_path):
                        for file in files:
                            full_path = os.path.join(root, file)
                            # ZIP 내부 경로: 주주총회의사록 및 조정산식/4130원/파일명.pdf
                            arc_path = os.path.relpath(full_path, output_dir)
                            zf.write(full_path, arc_path)

    buf.seek(0)

    # ZIP 파일명: Step03_납입서류_회차명_날짜.zip
    today = datetime.now().strftime('%Y%m%d')
    zip_filename = f"Step03_납입서류_{round_obj['name']}_{today}.zip"

    return send_file(
        buf,
        mimetype='application/zip',
        as_attachment=True,
        download_name=zip_filename
    )


# ── Step 03-3 (의무보유) routes ────────────────────────────────────────────────

@app.route('/round/<int:round_id>/step033')
def step033(round_id):
    round_obj = db.get_round(round_id)
    if not round_obj:
        abort(404)
    config = db.get_holding_config(round_id)
    subjects = db.get_holding_subjects(round_id)
    outputs = db.get_step_outputs(round_id, 'step033')
    return render_template(
        'step033.html',
        round=round_obj,
        config=config,
        subjects=subjects,
        outputs=outputs,
    )


@app.route('/round/<int:round_id>/step033/config', methods=['POST'])
def step033_config(round_id):
    data = request.get_json(silent=True) or {}
    db.save_holding_config(
        round_id,
        data.get('holding_start', ''),
        data.get('holding_end', ''),
        data.get('doc_date', ''),
        data.get('processing_date', ''),
    )
    return jsonify(success=True)


@app.route('/round/<int:round_id>/step033/subjects', methods=['GET'])
def step033_subjects_list(round_id):
    subjects = db.get_holding_subjects(round_id)
    return jsonify(success=True, data=subjects)


@app.route('/round/<int:round_id>/step033/subjects/add', methods=['POST'])
def step033_subject_add(round_id):
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify(success=False, message='이름을 입력하세요.')
    try:
        qty = int(data.get('quantity') or 0)
    except (ValueError, TypeError):
        qty = 0
    db.add_holding_subject(
        round_id, name,
        (data.get('relationship') or '미등기임원').strip(),
        qty,
        (data.get('branch') or '도곡').strip(),
        (data.get('account_number') or '').strip(),
        (data.get('note') or '').strip(),
    )
    subjects = db.get_holding_subjects(round_id)
    return jsonify(success=True, data=subjects)


@app.route('/round/<int:round_id>/step033/subjects/<int:subject_id>', methods=['PUT'])
def step033_subject_update(round_id, subject_id):
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify(success=False, message='이름을 입력하세요.')
    try:
        qty = int(data.get('quantity') or 0)
    except (ValueError, TypeError):
        qty = 0
    db.update_holding_subject(
        subject_id, name,
        (data.get('relationship') or '미등기임원').strip(),
        qty,
        (data.get('branch') or '도곡').strip(),
        (data.get('account_number') or '').strip(),
        (data.get('note') or '').strip(),
    )
    subjects = db.get_holding_subjects(round_id)
    return jsonify(success=True, data=subjects)


@app.route('/round/<int:round_id>/step033/subjects/<int:subject_id>', methods=['DELETE'])
def step033_subject_delete(round_id, subject_id):
    db.delete_holding_subject(subject_id)
    subjects = db.get_holding_subjects(round_id)
    return jsonify(success=True, data=subjects)


@app.route('/round/<int:round_id>/step033/applicants-list', methods=['GET'])
def step033_applicants_list(round_id):
    """신청자 목록을 이름별 합산하여 반환 (선택 모달용)."""
    applicants = db.get_applicants(round_id)
    from collections import OrderedDict
    person_map = OrderedDict()
    for ap in applicants:
        name = ap['name']
        if name not in person_map:
            person_map[name] = {
                'name': name,
                'quantity': 0,
                'account_number': ap.get('account_number') or '',
                'relationship': '미등기임원',
            }
        person_map[name]['quantity'] += (ap.get('quantity') or 0)
    return jsonify(success=True, data=list(person_map.values()))


@app.route('/round/<int:round_id>/step033/load-applicants', methods=['POST'])
def step033_load_applicants(round_id):
    """선택된 신청자들을 의무보유 대상자로 추가.
    data.selected: [{name, quantity, account_number, relationship}, ...]
    data.mode: 'append' | 'replace'
    """
    data = request.get_json(silent=True) or {}
    selected = data.get('selected', [])
    mode = data.get('mode', 'append')

    if not selected:
        return jsonify(success=False, message='선택된 인원이 없습니다.')

    if mode == 'replace':
        db.delete_all_holding_subjects(round_id)

    # 이미 있는 이름 목록 (append 시 중복 방지)
    existing = {s['name'] for s in db.get_holding_subjects(round_id)}

    added = 0
    for person in selected:
        name = (person.get('name') or '').strip()
        if not name:
            continue
        if mode == 'append' and name in existing:
            continue
        try:
            qty = int(person.get('quantity') or 0)
        except (ValueError, TypeError):
            qty = 0
        db.add_holding_subject(
            round_id, name,
            (person.get('relationship') or '미등기임원').strip(),
            qty,
            '도곡',
            (person.get('account_number') or '').strip(),
            '',
        )
        added += 1

    subjects = db.get_holding_subjects(round_id)
    return jsonify(success=True, data=subjects, added=added, count=len(subjects))


@app.route('/round/<int:round_id>/step033/generate', methods=['POST'])
def step033_generate(round_id):
    round_obj = db.get_round(round_id)
    if not round_obj:
        return jsonify(success=False, message='회차를 찾을 수 없습니다.'), 404

    data = request.get_json(silent=True) or {}
    doc_types = data.get('doc_types', ['hwakjakseo', 'gongmun'])

    config = db.get_holding_config(round_id)
    subjects = db.get_holding_subjects(round_id)

    if not subjects:
        return jsonify(success=False, message='의무보유 대상자가 없습니다.')

    holding_start = config.get('holding_start', '')
    holding_end   = config.get('holding_end', '')
    doc_date      = config.get('doc_date', '')
    proc_date     = config.get('processing_date', '')

    output_dir = os.path.join(OUTPUT_FOLDER, str(round_id), 'step033')
    os.makedirs(output_dir, exist_ok=True)

    results = []

    # 1. 의무보유확약서
    if 'hwakjakseo' in doc_types:
        try:
            tpl = os.path.join(TEMPLATES_HWP, 'hwuboyu_hwakjakseo_template.docx')
            out = os.path.join(output_dir, 'hwakjakseo.docx')
            subj_list = [
                {
                    'name': s['name'],
                    'relationship': s['relationship'],
                    'quantity': s['quantity'],
                    'note': s['note'] or '',
                }
                for s in subjects
            ]
            generate_hwakjakseo(tpl, out, subj_list, holding_start, holding_end)
            db.save_step_output(round_id, 'step033', 'hwakjakseo.docx', out)
            results.append({'name': '의무보유확약서.docx', 'filename': 'hwakjakseo.docx', 'success': True})
        except Exception as e:
            results.append({'name': '의무보유확약서.docx', 'success': False, 'message': str(e)})

    # 2. 계속보유신청 공문
    if 'gongmun' in doc_types:
        try:
            tpl = os.path.join(TEMPLATES_HWP, 'kesokyuboyu_gongmun_template.docx')
            out = os.path.join(output_dir, 'gongmun.docx')

            # 이름별 계좌 목록 구성 (applicants에서) — 다중 계좌 지원
            from collections import defaultdict
            all_applicants = db.get_applicants(round_id)
            applicant_accounts = defaultdict(list)
            for ap in all_applicants:
                applicant_accounts[ap['name']].append({
                    'account_number': ap.get('account_number') or '',
                    'quantity': ap.get('quantity') or 0,
                })

            subjects_dogok   = [
                {
                    'name': s['name'],
                    'account_number': s['account_number'] or '',
                    'quantity': s['quantity'],
                    'branch': s['branch'],
                    'note': s['note'] or '주1)',
                }
                for s in subjects if s['branch'] == '도곡'
            ]
            subjects_yeouido = [
                {
                    'name': s['name'],
                    'account_number': s['account_number'] or '',
                    'quantity': s['quantity'],
                    'branch': s['branch'],
                    'note': s['note'] or '주1)',
                }
                for s in subjects if s['branch'] == '여의도'
            ]
            generate_gongmun(
                tpl, out, doc_date, proc_date,
                subjects_dogok, subjects_yeouido,
                holding_start, holding_end,
                applicant_accounts=dict(applicant_accounts),
            )
            db.save_step_output(round_id, 'step033', 'gongmun.docx', out)
            results.append({'name': '계속보유신청공문.docx', 'filename': 'gongmun.docx', 'success': True})
        except Exception as e:
            results.append({'name': '계속보유신청공문.docx', 'success': False, 'message': str(e)})

    return jsonify(success=True, data=results)


@app.route('/round/<int:round_id>/step033/download/<path:filename>')
def download_step033(round_id, filename):
    output_dir = os.path.join(OUTPUT_FOLDER, str(round_id), 'step033')
    safe = os.path.basename(filename)
    full_path = os.path.join(output_dir, safe)
    if not os.path.isfile(full_path):
        abort(404)
    display_map = {
        'hwakjakseo.docx': '의무보유확약서.docx',
        'gongmun.docx':    '계속보유신청공문.docx',
    }
    display = display_map.get(safe, safe)
    mime = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    return send_file(full_path, mimetype=mime, as_attachment=True, download_name=display)


# ── Step 04 (등기신청) routes ──────────────────────────────────────────────────

@app.route('/round/<int:round_id>/step04')
def step04(round_id):
    """Step 04 - 등기신청"""
    round_obj = db.get_round(round_id)
    if not round_obj:
        abort(404)

    applicants = db.get_applicants(round_id)
    outputs = db.get_step_outputs(round_id, 'step04')

    # 총 행사주식수 계산
    total_shares = sum(ap.get('quantity') or 0 for ap in applicants)

    return render_template(
        'step04.html',
        round=round_obj,
        applicants=applicants,
        outputs=outputs,
        total_shares=total_shares,
    )


@app.route('/round/<int:round_id>/step04/generate', methods=['POST'])
def step04_generate(round_id):
    """Step 04 서류 생성"""
    round_obj = db.get_round(round_id)
    if not round_obj:
        return jsonify(success=False, message='회차를 찾을 수 없습니다.'), 404

    applicants = db.get_applicants(round_id)
    if not applicants:
        return jsonify(success=False, message='신청자 데이터가 없습니다.')

    output_dir = os.path.join(OUTPUT_FOLDER, str(round_id), 'step04')
    templates_dir = os.path.join(BASE_DIR, 'templates_step04')

    try:
        results = generate_step04_documents(round_id, applicants, output_dir, templates_dir)

        # 생성된 파일들 DB에 저장
        for file_info in results.get('files', []):
            filename = os.path.basename(file_info['path'])
            db.save_step_output(round_id, 'step04', filename, file_info['path'])

        return jsonify(success=True, data=results)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(success=False, message=str(e))


@app.route('/round/<int:round_id>/download_step04/<path:filename>')
def download_step04(round_id, filename):
    """Step 04 서류 다운로드"""
    output_dir = os.path.join(OUTPUT_FOLDER, str(round_id), 'step04')
    safe = os.path.basename(filename)
    full_path = os.path.join(output_dir, safe)

    if not os.path.isfile(full_path):
        abort(404)

    # MIME 타입 자동 감지
    if safe.endswith('.pdf'):
        mime = 'application/pdf'
    elif safe.endswith('.hwpx'):
        mime = 'application/x-hwpx'
    elif safe.endswith('.xlsx'):
        mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    else:
        mime = 'application/octet-stream'

    return send_file(full_path, mimetype=mime, as_attachment=True, download_name=safe)


# ── Step 05 – 예탁원 신주발행의뢰 ─────────────────────────────────────────────

@app.route('/round/<int:round_id>/step05')
def step05(round_id):
    round_obj = db.get_round(round_id)
    if not round_obj:
        abort(404)
    applicants = db.get_applicants(round_id)
    prices     = db.get_prices_for_round(round_id)
    config     = db.get_issuance_config(round_id)
    outputs    = db.get_step_outputs(round_id, 'step05')

    # 행사가액별 신청자 분류
    price_groups = {}
    for p in prices:
        group = [ap for ap in applicants if ap.get('exercise_price') == p]
        total_qty = sum(ap.get('quantity') or 0 for ap in group)
        price_groups[p] = {'applicants': group, 'count': len(group), 'qty': total_qty}

    # 붙임8 업로드 현황
    attachment8_map = db.get_all_attachment8(round_id)

    return render_template(
        'step05.html',
        round=round_obj,
        applicants=applicants,
        prices=prices,
        config=config,
        outputs=outputs,
        price_groups=price_groups,
        attachment8_map=attachment8_map,
    )


@app.route('/round/<int:round_id>/step05/config', methods=['POST'])
def step05_config(round_id):
    data = request.get_json(silent=True) or {}
    try:
        db.save_issuance_config(
            round_id,
            data.get('payment_date', ''),
            data.get('dividend_base_date', ''),
            data.get('listing_date', ''),
            data.get('contact_name', '정민우'),
            data.get('contact_phone', '010-3615-4909'),
            data.get('stock_code', '488280'),
            data.get('agent_name', ''),
            data.get('agent_phone', ''),
            data.get('agent_rrn', ''),
            data.get('agent_address', '경기도 성남시 분당구 대왕판교로192번길 12, 3층'),
        )
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, message=str(e))


@app.route('/round/<int:round_id>/step05/generate', methods=['POST'])
def step05_generate(round_id):
    """발행가액별 신주발행의뢰 ZIP 파일 생성 (OCR 캐싱 적용)."""
    print("\n" + "="*80)
    print(f"[Step05 생성 시작] 회차 ID: {round_id}")
    print("="*80)

    round_obj = db.get_round(round_id)
    if not round_obj:
        return jsonify(success=False, message='회차를 찾을 수 없습니다.'), 404

    applicants = db.get_applicants(round_id)
    if not applicants:
        return jsonify(success=False, message='신청자 데이터가 없습니다.')

    prices = db.get_prices_for_round(round_id)
    if not prices:
        return jsonify(success=False, message='행사가액이 등록되지 않았습니다.')

    print(f"\n[데이터 로드 완료]")
    print(f"  - 신청자: {len(applicants)}명")
    print(f"  - 행사가액: {len(prices)}개 ({', '.join(f'{p:,}원' for p in prices)})")

    config = db.get_issuance_config(round_id)
    attachment8_map = db.get_all_attachment8(round_id)

    # ─────────────────────────────────────────────────────────────
    # OCR 캐싱: 이미 추출된 데이터는 DB에서 가져오고, 없으면 OCR 실행
    # 메모리 절약을 위해 소규모 배치로 처리
    # ─────────────────────────────────────────────────────────────
    import gc
    from processors.ocr_reader import extract_account_and_broker

    print(f"\n[OCR 데이터 준비 중...]")
    ocr_rrn_success = 0
    ocr_rrn_fail = 0
    ocr_account_success = 0
    ocr_account_fail = 0

    # 배치 크기 설정 (메모리 절약을 위해 작은 배치로 처리)
    BATCH_SIZE = 3
    total_applicants = len(applicants)

    for batch_start in range(0, total_applicants, BATCH_SIZE):
        batch_end = min(batch_start + BATCH_SIZE, total_applicants)
        batch = applicants[batch_start:batch_end]

        print(f"\n  ── 배치 {batch_start+1}~{batch_end}/{total_applicants} 처리 중 ──")

        for ap in batch:
            ap_id = ap['id']
            ap_name = ap.get('name', '?')

            # 주민번호: DB에 없으면 OCR
            if not ap.get('rrn'):
                id_docs = db.get_documents_for_applicant_ids([ap_id], 'id_copy')
                if id_docs:
                    try:
                        rrn_map = extract_rrn_batch(id_docs)
                        rrn = rrn_map.get(ap_id)
                        if rrn:
                            db.update_applicant_ocr(ap_id, rrn=rrn)
                            ap['rrn'] = rrn
                            ocr_rrn_success += 1
                            print(f"  [OCR] {ap_name}: 주민번호 추출 성공")
                        else:
                            ocr_rrn_fail += 1
                            print(f"  [OCR] {ap_name}: 주민번호 추출 실패")
                    except Exception as e:
                        ocr_rrn_fail += 1
                        print(f"  [OCR] {ap_name}: 주민번호 추출 오류 - {e}")
            else:
                print(f"  [캐시] {ap_name}: 주민번호 이미 있음 ({ap.get('rrn')})")

            # 계좌번호/증권사: DB에 없으면 OCR
            if not ap.get('ocr_account'):
                acct_docs = db.get_documents_for_applicant_ids([ap_id], 'account_copy')
                if acct_docs:
                    try:
                        result = extract_account_and_broker(acct_docs[0]['file_path'])
                        account = result.get('account', '')
                        broker = result.get('broker', '')
                        if account or broker:
                            db.update_applicant_ocr(ap_id, ocr_account=account, broker=broker)
                            ap['ocr_account'] = account
                            ap['broker'] = broker
                            ocr_account_success += 1
                            print(f"  [OCR] {ap_name}: 계좌번호/증권사 추출 성공 ({broker} / {account})")
                        else:
                            ocr_account_fail += 1
                            print(f"  [OCR] {ap_name}: 계좌번호/증권사 추출 실패")
                    except Exception as e:
                        ocr_account_fail += 1
                        print(f"  [OCR] {ap_name}: 계좌번호/증권사 추출 오류 - {e}")
            else:
                print(f"  [캐시] {ap_name}: 계좌정보 이미 있음 ({ap.get('broker')} / {ap.get('ocr_account')})")

        # 배치 처리 후 메모리 정리
        if batch_end < total_applicants:
            print(f"  [메모리 정리 중...]")
            gc.collect()

    print(f"\n[OCR 요약]")
    print(f"  - 주민번호: 성공 {ocr_rrn_success}명, 실패 {ocr_rrn_fail}명")
    print(f"  - 계좌정보: 성공 {ocr_account_success}명, 실패 {ocr_account_fail}명")

    # OCR 완료 후 메모리 정리
    print(f"\n[OCR 완료 - 메모리 정리 중...]")
    gc.collect()

    # ─────────────────────────────────────────────────────────────
    # 발행가액별 ZIP 생성
    # ─────────────────────────────────────────────────────────────
    from processors.step05_generator import generate_step05_zip

    output_base = os.path.join(OUTPUT_FOLDER, str(round_id), 'step05')
    os.makedirs(output_base, exist_ok=True)

    print(f"\n[발행가액별 ZIP 생성 시작]")
    print(f"  출력 경로: {output_base}")

    results = []
    for price in prices:
        print(f"\n{'─'*80}")
        print(f"[{price:,}원 폴더 생성 중...]")
        print(f"{'─'*80}")

        group = [ap for ap in applicants if ap.get('exercise_price') == price]
        if not group:
            print(f"  WARNING 신청자 없음, 건너뜀")
            continue

        print(f"  신청자: {len(group)}명 ({', '.join(ap.get('name', '?') for ap in group[:5])}" +
              (f" 외 {len(group)-5}명" if len(group) > 5 else "") + ")")

        # 붙임8 파일 경로
        attachment8_file = None
        if price in attachment8_map:
            attachment8_file = attachment8_map[price].get('file_path')
            print(f"  붙임8: {attachment8_map[price].get('original_name')} OK")
        else:
            print(f"  붙임8: 업로드 안됨 (건너뜀)")

        try:
            result = generate_step05_zip(
                round_obj, group, price, config, attachment8_file, output_base
            )

            if result['success']:
                # ZIP 파일 정보 저장
                db.save_step_output(
                    round_id, 'step05', result['zip_name'], result['zip_path']
                )
                print(f"\n  OK 생성 완료: {result['zip_name']}")
                print(f"  - 포함 파일: {len(result['files'])}개")
                if result.get('errors'):
                    print(f"  - 경고: {len(result['errors'])}개")
                    for err in result['errors']:
                        print(f"    · {err}")

                results.append({
                    'success': True,
                    'name': f'{price:,}원 신주발행의뢰',
                    'filename': result['zip_name'],
                    'files_count': len(result['files']),
                    'message': result['message']
                })
            else:
                print(f"\n  FAIL 생성 실패: {result.get('message', '알 수 없는 오류')}")
                results.append({
                    'success': False,
                    'name': f'{price:,}원',
                    'message': result.get('message', '생성 실패')
                })
        except Exception as e:
            print(f"\n  FAIL 예외 발생: {e}")
            import traceback
            traceback.print_exc()
            results.append({
                'success': False,
                'name': f'{price:,}원',
                'message': str(e)
            })

    print(f"\n{'='*80}")
    print(f"[Step05 생성 완료] 성공: {sum(1 for r in results if r['success'])}개, 실패: {sum(1 for r in results if not r['success'])}개")
    print(f"{'='*80}\n")

    if not results:
        return jsonify(success=False, message='생성할 데이터가 없습니다.')

    return jsonify(success=True, data=results)


@app.route('/round/<int:round_id>/step05/download/<path:filename>')
def download_step05(round_id, filename):
    output_dir = os.path.join(OUTPUT_FOLDER, str(round_id), 'step05')
    safe       = filename.replace('\\', '/')
    full_path  = os.path.join(output_dir, safe)
    full_path  = os.path.normpath(full_path)
    # 보안: output_dir 밖으로 나가지 못하게
    if not full_path.startswith(os.path.normpath(output_dir)):
        abort(404)
    if not os.path.isfile(full_path):
        abort(404)
    display = os.path.basename(safe)
    if display.endswith('.xlsx'):
        mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    elif display.endswith('.zip'):
        mime = 'application/zip'
    else:
        mime = 'application/pdf'
    return send_file(full_path, mimetype=mime, as_attachment=True, download_name=display)


@app.route('/round/<int:round_id>/step05/upload_attachment8/<int:price>', methods=['POST'])
def upload_attachment8(round_id, price):
    """발행가액별 붙임8 (주식납입금보관증명서) 업로드."""
    if 'file' not in request.files:
        return jsonify(success=False, message='파일이 없습니다.')

    file = request.files['file']
    if not file or not file.filename:
        return jsonify(success=False, message='파일을 선택하세요.')

    # 허용 확장자: PDF, JPG, PNG
    allowed = {'.pdf', '.jpg', '.jpeg', '.png'}
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed:
        return jsonify(success=False, message='PDF 또는 이미지 파일만 업로드 가능합니다.')

    try:
        # 저장 경로
        upload_dir = os.path.join(UPLOAD_FOLDER, str(round_id), 'attachment8')
        os.makedirs(upload_dir, exist_ok=True)

        # 파일명: attachment8_<가액><확장자>
        file_name = f'attachment8_{price}{ext}'
        file_path = os.path.join(upload_dir, file_name)

        # 기존 파일 삭제
        if os.path.isfile(file_path):
            os.remove(file_path)

        # 저장
        file.save(file_path)

        # DB에 저장
        db.save_attachment8(round_id, price, file_name, file.filename, file_path)

        return jsonify(success=True, message='업로드 완료', filename=file.filename)
    except Exception as e:
        return jsonify(success=False, message=str(e))


@app.route('/round/<int:round_id>/step05/delete_attachment8/<int:price>', methods=['POST'])
def delete_attachment8_route(round_id, price):
    """발행가액별 붙임8 삭제."""
    try:
        attachment = db.get_attachment8(round_id, price)
        if attachment:
            # 파일 삭제
            if os.path.isfile(attachment['file_path']):
                os.remove(attachment['file_path'])
            # DB 삭제
            db.delete_attachment8(round_id, price)
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, message=str(e))


# ── Employee self-submit routes ────────────────────────────────────────────────

@app.route('/submit/<token>', methods=['GET'])
def employee_submit(token):
    ap = db.get_applicant_by_token(token)
    if not ap:
        abort(404)
    round_obj = db.get_round(ap['round_id'])
    docs = db.get_documents(ap['id'])
    doc_map = {d['doc_type']: d for d in docs}
    return render_template(
        'employee_submit.html',
        applicant=ap,
        round=round_obj,
        doc_map=doc_map,
        submitted=(len(doc_map) >= 3),
    )


@app.route('/submit/<token>', methods=['POST'])
def employee_submit_post(token):
    ap = db.get_applicant_by_token(token)
    if not ap:
        abort(404)

    saved = []
    errors = []

    for doc_type in ['application', 'id_copy', 'account_copy']:
        file = request.files.get(doc_type)
        if file and file.filename and allowed_file(file.filename):
            unique_name, file_path, original = save_uploaded_file(
                file, ap['round_id'], ap['id'], doc_type
            )
            db.add_document(ap['id'], doc_type, unique_name, original, file_path)
            saved.append(DOC_TYPE_LABELS[doc_type])
        elif file and file.filename and not allowed_file(file.filename):
            errors.append(f"{DOC_TYPE_LABELS.get(doc_type, doc_type)}: 허용되지 않는 파일 형식")

    docs = db.get_documents(ap['id'])
    doc_map = {d['doc_type']: d for d in docs}
    all_submitted = len(doc_map) >= 3

    round_obj = db.get_round(ap['round_id'])
    return render_template(
        'employee_submit.html',
        applicant=ap,
        round=round_obj,
        doc_map=doc_map,
        submitted=all_submitted,
        saved=saved,
        errors=errors,
        just_submitted=True,
    )


# ── Error handlers ─────────────────────────────────────────────────────────────

# ── STEP 06: KIND 상장신청 ────────────────────────────────────────────────────

@app.route('/round/<int:round_id>/step06')
def step06(round_id):
    """Step06 KIND 상장신청 페이지."""
    round_obj = db.get_round(round_id)
    if not round_obj:
        abort(404)

    # 설정 조회
    config = db.get_step06_config(round_id)

    # 발행가액 목록
    exercise_prices = db.get_exercise_prices(round_id)

    # 발행등록사실확인서 업로드 현황
    issuance_confirmations = db.get_step06_issuance_confirmations(round_id)

    # 생성 이력
    outputs = db.get_step_outputs(round_id, 'step06')

    return render_template(
        'step06.html',
        round=round_obj,
        config=config,
        exercise_prices=[p['price'] for p in exercise_prices],
        issuance_confirmations=issuance_confirmations,
        outputs=outputs
    )


@app.route('/api/round/<int:round_id>/step06/config', methods=['POST'])
def step06_config(round_id):
    """Step06 기본 설정 저장."""
    data = request.json
    db.save_step06_config(
        round_id,
        submission_date=data.get('submission_date')
    )
    return jsonify(success=True)


@app.route('/api/round/<int:round_id>/step06/upload/listing_fee', methods=['POST'])
def upload_listing_fee(round_id):
    """상장수수료 납부영수증 업로드."""
    if 'file' not in request.files:
        return jsonify(success=False, error='파일이 없습니다.')

    file = request.files['file']
    if not file or not file.filename:
        return jsonify(success=False, error='파일을 선택하세요.')

    try:
        upload_dir = os.path.join(UPLOAD_FOLDER, str(round_id), 'step06')
        os.makedirs(upload_dir, exist_ok=True)

        filename = secure_filename(file.filename)
        file_path = os.path.join(upload_dir, f'listing_fee_{filename}')
        file.save(file_path)

        db.save_step06_config(round_id, listing_fee_receipt=file_path)

        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e))


@app.route('/api/round/<int:round_id>/step06/upload/holding_proof', methods=['POST'])
def upload_holding_proof(round_id):
    """의무보유증명서 및 의무보유청구내역 폴더 업로드."""
    files = request.files.getlist('files')
    if not files:
        return jsonify(success=False, error='파일이 없습니다.')

    try:
        upload_dir = os.path.join(UPLOAD_FOLDER, str(round_id), 'step06', 'holding_proof')
        os.makedirs(upload_dir, exist_ok=True)

        for file in files:
            if file and file.filename:
                filename = secure_filename(file.filename)
                file_path = os.path.join(upload_dir, filename)
                os.makedirs(os.path.dirname(file_path), exist_ok=True)
                file.save(file_path)

        db.save_step06_config(round_id, holding_proof_folder=upload_dir)

        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e))


@app.route('/api/round/<int:round_id>/step06/upload/issuance_confirmation', methods=['POST'])
def upload_issuance_confirmation(round_id):
    """발행가액별 발행등록사실확인서 업로드."""
    if 'file' not in request.files:
        return jsonify(success=False, error='파일이 없습니다.')

    file = request.files['file']
    price = request.form.get('exercise_price')

    if not file or not file.filename or not price:
        return jsonify(success=False, error='파일과 발행가액을 모두 입력하세요.')

    try:
        price = int(price)
        upload_dir = os.path.join(UPLOAD_FOLDER, str(round_id), 'step06', 'issuance_confirmations')
        os.makedirs(upload_dir, exist_ok=True)

        filename = secure_filename(file.filename)
        file_path = os.path.join(upload_dir, f'{price}_{filename}')
        file.save(file_path)

        db.save_step06_issuance_confirmation(round_id, price, file_path, filename)

        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e))


@app.route('/api/round/<int:round_id>/step06/upload/issuance_folder', methods=['POST'])
def upload_issuance_folder(round_id):
    """발행등록사실확인서 폴더 업로드 (여러 파일 자동 매칭)."""
    files = request.files.getlist('files')
    if not files:
        return jsonify(success=False, error='파일이 없습니다.')

    try:
        upload_dir = os.path.join(UPLOAD_FOLDER, str(round_id), 'step06', 'issuance_confirmations')
        os.makedirs(upload_dir, exist_ok=True)

        exercise_prices = [p['price'] for p in db.get_exercise_prices(round_id)]
        uploaded_count = 0

        for file in files:
            if file and file.filename:
                filename = secure_filename(file.filename)

                # 파일명에서 발행가액 추출 시도
                matched_price = None
                for price in exercise_prices:
                    if str(price) in filename:
                        matched_price = price
                        break

                file_path = os.path.join(upload_dir, filename)
                file.save(file_path)

                if matched_price:
                    db.save_step06_issuance_confirmation(round_id, matched_price, file_path, filename)
                    uploaded_count += 1

        return jsonify(success=True, uploaded_count=uploaded_count)
    except Exception as e:
        return jsonify(success=False, error=str(e))


@app.route('/api/round/<int:round_id>/step06/upload/employment_cert', methods=['POST'])
def upload_employment_cert(round_id):
    """재직증명서 폴더 업로드."""
    files = request.files.getlist('files')
    if not files:
        return jsonify(success=False, error='파일이 없습니다.')

    try:
        upload_dir = os.path.join(UPLOAD_FOLDER, str(round_id), 'step06', 'employment_cert')
        os.makedirs(upload_dir, exist_ok=True)

        for file in files:
            if file and file.filename:
                filename = secure_filename(file.filename)
                file_path = os.path.join(upload_dir, filename)
                os.makedirs(os.path.dirname(file_path), exist_ok=True)
                file.save(file_path)

        db.save_step06_config(round_id, employment_cert_folder=upload_dir)

        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e))


@app.route('/api/round/<int:round_id>/step06/upload/exercise_summary', methods=['POST'])
def upload_exercise_summary(round_id):
    """스톡옵션 행사현황표 엑셀 업로드."""
    if 'file' not in request.files:
        return jsonify(success=False, error='파일이 없습니다.')

    file = request.files['file']
    if not file or not file.filename:
        return jsonify(success=False, error='파일을 선택하세요.')

    try:
        upload_dir = os.path.join(UPLOAD_FOLDER, str(round_id), 'step06')
        os.makedirs(upload_dir, exist_ok=True)

        filename = secure_filename(file.filename)
        file_path = os.path.join(upload_dir, f'exercise_summary_{filename}')
        file.save(file_path)

        db.save_step06_config(round_id, exercise_summary_excel=file_path)

        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e))


@app.route('/api/round/<int:round_id>/step06/generate', methods=['POST'])
def step06_generate(round_id):
    """KIND 상장신청 ZIP 파일 생성."""
    round_obj = db.get_round(round_id)
    if not round_obj:
        return jsonify(success=False, message='회차를 찾을 수 없습니다.')

    config = db.get_step06_config(round_id)
    output_base = os.path.join(OUTPUT_FOLDER, str(round_id))

    try:
        result = generate_step06_zip(round_obj, config, output_base)

        if result['success']:
            # 생성 이력 저장
            zip_filename = os.path.basename(result['zip_path'])
            rel_path = os.path.relpath(result['zip_path'], OUTPUT_FOLDER)
            db.save_step_output(round_id, 'step06', zip_filename, rel_path)

            return jsonify(
                success=True,
                message=result['message'],
                warnings=result.get('warnings', [])
            )
        else:
            return jsonify(success=False, error=result['message'])

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(success=False, error=str(e))


@app.errorhandler(404)
def not_found(e):
    return render_template('404.html'), 404


@app.errorhandler(413)
def too_large(e):
    return jsonify(success=False, message='파일 크기가 너무 큽니다. (최대 50MB)'), 413


# ── Startup ────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    db.init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)
