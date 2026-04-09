"""신분증 OCR로 주민등록번호 추출."""
import re
import os
import io
from PIL import Image

_reader = None


def _get_reader():
    """EasyOCR reader 싱글턴 (첫 호출 시 모델 로드)."""
    global _reader
    if _reader is None:
        import easyocr
        _reader = easyocr.Reader(['ko', 'en'], gpu=False, verbose=False)
    return _reader


# 주민등록번호 패턴: 6자리-7자리
_RRN_PATTERN = re.compile(r'(\d{6})\s*[-–—]\s*(\d{7})')
# 느슨한 패턴 (OCR 오류 대비: 구분자가 . 이나 공백인 경우)
_RRN_LOOSE = re.compile(r'(\d{6})\s*[-–—.,\s]\s*(\d{6,7})')

# 계좌번호 패턴들
# 일반 형식: 123-456789-01, 12-34-567890, 123-45-678901 등
_ACCOUNT_PATTERNS = [
    re.compile(r'(\d{2,4})\s*[-–—]\s*(\d{4,8})\s*[-–—]\s*(\d{2,4})'),  # 3분할: 123-456789-01
    re.compile(r'(\d{2,4})\s*[-–—]\s*(\d{6,12})'),  # 2분할: 123-456789012
    re.compile(r'\b(\d{10,14})\b'),  # 숫자만: 12345678901234
]


def _images_from_pdf(pdf_path):
    """pypdf로 PDF에서 모든 이미지를 추출."""
    images = []
    try:
        import pypdf
        reader = pypdf.PdfReader(pdf_path)
        for page in reader.pages:
            if '/XObject' not in (page.get('/Resources') or {}):
                continue
            x_objects = page['/Resources']['/XObject'].get_object()
            for obj_name in x_objects:
                obj = x_objects[obj_name].get_object()
                if obj.get('/Subtype') != '/Image':
                    continue
                data = obj.get_data()
                width = obj.get('/Width', 0)
                height = obj.get('/Height', 0)
                filt = obj.get('/Filter')
                # 배열 형태 필터 처리 (예: ['/FlateDecode', '/DCTDecode'])
                if isinstance(filt, list):
                    filt = filt[-1] if filt else None
                elif hasattr(filt, 'get_object'):
                    filt = str(filt.get_object())
                else:
                    filt = str(filt) if filt else None

                try:
                    if filt in ('/DCTDecode', '/JPXDecode'):
                        images.append(Image.open(io.BytesIO(data)))
                    elif width and height:
                        cs = str(obj.get('/ColorSpace', ''))
                        if 'RGB' in cs:
                            images.append(Image.frombytes('RGB', (width, height), data))
                        elif 'Gray' in cs:
                            images.append(Image.frombytes('L', (width, height), data))
                except Exception:
                    pass
    except Exception:
        pass
    return images


def _load_images(file_path):
    """파일 경로에서 이미지 목록 로드."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext in ('.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif'):
        try:
            return [Image.open(file_path)]
        except Exception:
            return []
    elif ext == '.pdf':
        return _images_from_pdf(file_path)
    return []


def _find_rrn_in_text(text):
    """텍스트에서 주민번호 패턴 찾기."""
    # 정확한 패턴
    m = _RRN_PATTERN.search(text)
    if m:
        return f'{m.group(1)}-{m.group(2)}'

    # 느슨한 패턴
    m = _RRN_LOOSE.search(text)
    if m:
        part2 = m.group(2)
        if len(part2) == 7:
            return f'{m.group(1)}-{part2}'
        elif len(part2) == 6:
            return f'{m.group(1)}-{part2}*'

    # 공백 제거 후 재시도
    text_no_space = text.replace(' ', '')
    m = _RRN_PATTERN.search(text_no_space)
    if m:
        return f'{m.group(1)}-{m.group(2)}'

    return ''


def extract_rrn(file_path):
    """
    신분증 이미지/PDF에서 주민등록번호를 추출.
    PDF는 모든 이미지를 추출하여 각각 OCR 시도.
    Returns: 'XXXXXX-XXXXXXX' 형태 문자열 or ''
    """
    if not os.path.isfile(file_path):
        return ''

    images = _load_images(file_path)
    if not images:
        return ''

    import numpy as np
    reader = _get_reader()

    # 크기가 큰 이미지부터 시도 (신분증 본문일 가능성 높음)
    images.sort(key=lambda img: img.size[0] * img.size[1], reverse=True)

    for img in images:
        try:
            w, h = img.size
            if w < 100 or h < 100:
                continue  # 너무 작은 이미지 스킵

            # 작은 이미지는 확대
            if w < 500:
                ratio = 1000 / w
                img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)

            img_array = np.array(img.convert('RGB'))
            results = reader.readtext(img_array, detail=0)
            full_text = ' '.join(results)

            rrn = _find_rrn_in_text(full_text)
            if rrn:
                return rrn
        except Exception:
            continue

    return ''


def _find_account_in_text(text):
    """텍스트에서 계좌번호 패턴 찾기."""
    for pattern in _ACCOUNT_PATTERNS:
        m = pattern.search(text)
        if m:
            # 3분할 패턴
            if len(m.groups()) == 3:
                return f'{m.group(1)}-{m.group(2)}-{m.group(3)}'
            # 2분할 패턴
            elif len(m.groups()) == 2:
                return f'{m.group(1)}-{m.group(2)}'
            # 숫자만
            else:
                return m.group(1)

    # 공백 제거 후 재시도
    text_no_space = text.replace(' ', '')
    for pattern in _ACCOUNT_PATTERNS:
        m = pattern.search(text_no_space)
        if m:
            if len(m.groups()) == 3:
                return f'{m.group(1)}-{m.group(2)}-{m.group(3)}'
            elif len(m.groups()) == 2:
                return f'{m.group(1)}-{m.group(2)}'
            else:
                return m.group(1)

    return ''


def _load_broker_keywords():
    """증권사 키워드 리스트 로드 (code.xlsx 또는 기본값)."""
    try:
        import openpyxl
        excel_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'code.xlsx')

        if os.path.isfile(excel_path):
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            ws = wb.active
            keywords = []

            for i in range(2, ws.max_row + 1):
                row = list(ws[i])
                if len(row) >= 2:
                    name = row[1].value
                    if name:
                        name_str = str(name).strip()
                        keywords.append(name_str)
                        # 짧은 버전도 추가 (예: "미래에셋증권" → "미래에셋")
                        if '증권' in name_str:
                            keywords.append(name_str.replace('증권', ''))

            wb.close()
            return keywords
    except Exception:
        pass

    # 기본 키워드
    return ['미래에셋', '삼성', 'NH투자', 'KB', '한국투자', '키움', '신한투자',
            '하나', '대신', '메리츠', '토스', '카카오페이', '한화투자']


_BROKER_KEYWORDS = None


# 영문 → 한글 증권사 매핑
_BROKER_ENGLISH_MAP = {
    'mirae': '미래에셋',
    'miraeasset': '미래에셋',
    'samsung': '삼성',
    'kb': 'KB',
    'nh': 'NH투자',
    'kiwoom': '키움',
    'shinhan': '신한투자',
    'hana': '하나',
    'dashin': '대신',
    'daishin': '대신',
    'meritz': '메리츠',
    'toss': '토스',
    'kakaopay': '카카오페이',
    'hanwha': '한화투자',
    'sk': 'SK',
    'db': 'DB',
    'yuanta': '유안타',
    'hi': '한국투자',
    'ebest': '이베스트',
    'ls': 'LS',
}


def _find_broker_in_text(text):
    """텍스트에서 증권사명 찾기 (한글 + 영문 지원)."""
    global _BROKER_KEYWORDS

    if _BROKER_KEYWORDS is None:
        _BROKER_KEYWORDS = _load_broker_keywords()

    # 텍스트 정리
    text_clean = text.replace(' ', '').replace('\n', '').replace('-', '')
    text_lower = text_clean.lower()

    # 1. 한글 키워드 매칭 (가장 긴 것부터)
    for keyword in sorted(_BROKER_KEYWORDS, key=len, reverse=True):
        keyword_clean = keyword.replace(' ', '').replace('-', '')
        if keyword_clean in text_clean:
            return keyword

    # 2. 영문 키워드 매칭
    for eng, kor in _BROKER_ENGLISH_MAP.items():
        if eng.lower() in text_lower:
            # 한글 키워드 중에서 매칭되는 것 찾기
            for keyword in _BROKER_KEYWORDS:
                if kor in keyword:
                    return keyword
            return kor

    return ''


def extract_account_number(file_path):
    """
    계좌사본 이미지/PDF에서 계좌번호를 추출.
    Returns: 계좌번호 문자열 or ''
    """
    result = extract_account_and_broker(file_path)
    return result.get('account', '')


def extract_account_and_broker(file_path):
    """
    계좌사본 이미지/PDF에서 계좌번호와 증권사명 추출.
    Returns: {'account': '123-456', 'broker': '미래에셋증권'}
    """
    if not os.path.isfile(file_path):
        return {'account': '', 'broker': ''}

    images = _load_images(file_path)
    if not images:
        return {'account': '', 'broker': ''}

    import numpy as np
    reader = _get_reader()

    # 크기가 큰 이미지부터 시도
    images.sort(key=lambda img: img.size[0] * img.size[1], reverse=True)

    for img in images:
        try:
            w, h = img.size
            if w < 100 or h < 100:
                continue

            # 작은 이미지는 확대
            if w < 500:
                ratio = 1000 / w
                img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)

            img_array = np.array(img.convert('RGB'))
            results = reader.readtext(img_array, detail=0)
            full_text = ' '.join(results)

            account = _find_account_in_text(full_text)
            broker = _find_broker_in_text(full_text)

            # 둘 중 하나라도 찾으면 반환
            if account or broker:
                return {'account': account, 'broker': broker}
        except Exception:
            continue

    return {'account': '', 'broker': ''}


def extract_rrn_batch(applicant_docs):
    """
    여러 신청자의 신분증에서 일괄 주민번호 추출.
    applicant_docs: list of {'applicant_id': int, 'name': str, 'file_path': str}
    Returns: {applicant_id: 'XXXXXX-XXXXXXX' or ''}
    """
    results = {}
    total = len(applicant_docs)
    for idx, doc in enumerate(applicant_docs, 1):
        aid = doc['applicant_id']
        name = doc.get('name', '?')
        print(f"    [{idx}/{total}] {name} 신분증 OCR 중...", end=' ')
        rrn = extract_rrn(doc['file_path'])
        if rrn:
            print(f"✓ {rrn}")
        else:
            print(f"✗ 실패")
        results[aid] = rrn
    return results


def extract_account_batch(applicant_docs, debug=False):
    """
    여러 신청자의 계좌사본에서 일괄 계좌번호 + 증권사명 추출.
    applicant_docs: list of {'applicant_id': int, 'name': str, 'file_path': str}
    Returns: {applicant_id: {'account': '...', 'broker': '...', 'ocr_text': '...'}}
    """
    results = {}
    total = len(applicant_docs)
    for idx, doc in enumerate(applicant_docs, 1):
        aid = doc['applicant_id']
        name = doc.get('name', '?')
        print(f"    [{idx}/{total}] {name} 계좌사본 OCR 중...", end=' ')

        # OCR 텍스트도 함께 반환하도록 수정
        result = extract_account_and_broker_with_text(doc['file_path'])
        account = result.get('account', '')
        broker = result.get('broker', '')
        ocr_text = result.get('ocr_text', '')

        if account or broker:
            parts = []
            if broker:
                parts.append(f"증권사:{broker}")
            if account:
                parts.append(f"계좌:{account}")
            print(f"✓ {', '.join(parts)}")
        else:
            print(f"✗ 실패")

        results[aid] = result
    return results


def extract_account_and_broker_with_text(file_path):
    """
    계좌사본에서 계좌번호, 증권사명, OCR 텍스트 추출.
    Returns: {'account': '...', 'broker': '...', 'ocr_text': '...'}
    """
    if not os.path.isfile(file_path):
        return {'account': '', 'broker': '', 'ocr_text': ''}

    images = _load_images(file_path)
    if not images:
        return {'account': '', 'broker': '', 'ocr_text': ''}

    import numpy as np
    reader = _get_reader()

    # 크기가 큰 이미지부터 시도
    images.sort(key=lambda img: img.size[0] * img.size[1], reverse=True)

    for img in images:
        try:
            w, h = img.size
            if w < 100 or h < 100:
                continue

            # 작은 이미지는 확대
            if w < 500:
                ratio = 1000 / w
                img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)

            img_array = np.array(img.convert('RGB'))
            results = reader.readtext(img_array, detail=0)
            full_text = ' '.join(results)

            account = _find_account_in_text(full_text)
            broker = _find_broker_in_text(full_text)

            # OCR 텍스트도 함께 반환
            return {'account': account, 'broker': broker, 'ocr_text': full_text}
        except Exception:
            continue

    return {'account': '', 'broker': '', 'ocr_text': ''}
