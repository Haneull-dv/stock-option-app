"""행사내역 / 등기신청 / 신주발행 세부내역 엑셀 생성."""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os


_THIN = Side(style='thin')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# 증권사 → 계좌관리번호 매핑 (한국예탁결제원 기준)
BROKER_CODE_MAP = {
    '교보증권': 1,
    '신한투자증권': 2, '신한': 2,
    '한국투자증권': 3, '한국투자': 3, '한투': 3,
    '대신증권': 4, '대신': 4,
    '미래에셋증권': 5, '미래에셋': 5,
    '신영증권': 6,
    '유진투자증권': 8, '유진': 8,
    '한양증권': 9,
    '메리츠증권': 10, '메리츠': 10,
    'NH투자증권': 12, 'NH': 12,
    '부국증권': 13,
    'KB증권': 17, 'KB': 17,
    '한화투자증권': 21, '한화': 21,
    '현대차증권': 22,
    '유화증권': 23,
    '유안타증권': 24, '유안타': 24,
    'SK증권': 25,
    '상상인증권': 29,
    '삼성증권': 30, '삼성': 30,
    'DB금융투자': 31, 'DB': 31,
    '아이엠증권': 46,
    '키움증권': 50, '키움': 50,
    '리딩투자증권': 52,
    '하나증권': 56, '하나': 56,
    '아이비케이투자증권': 68, 'IBK투자증권': 68,
    '카카오페이증권': 69, '카카오페이': 69,
    '디에스투자증권': 70,
    '다올투자증권': 71, '다올': 71,
    '케이프투자증권': 72,
    '한국스탠다드차타드증권': 74,
    '토스증권': 77, '토스': 77,
    '비엔케이투자증권': 86, 'BNK투자증권': 86,
    'LS증권': 752,
    '코리아에셋투자증권': 753,
    '한국증권금융': 800,
    '도이치은행': 850,
    '한국씨티은행': 1480,
    '한국스탠다드차타드은행': 1500,
    '홍콩상하이은행': 1520,
}


def _cell(ws, row, col, value, bold=False, align='left', number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name='맑은 고딕', bold=bold, size=10)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    c.border = _BORDER
    if number_format:
        c.number_format = number_format
    return c


def generate_exercise_excel(round_name: str, exercise_date: str,
                             applicants: list, output_path: str) -> str:
    """
    행사내역 엑셀 생성.
    applicants: list of dicts with name, exercise_price, quantity, broker, account_number
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '행사내역'

    # 제목
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = f'스톡옵션 행사내역 ({exercise_date})'
    title_cell.font = Font(name='맑은 고딕', bold=True, size=13)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # 헤더
    headers = ['이름', '행사가액(원)', '수량(주)', '납입금액(원)', '증권사', '계좌번호', '비고']
    for ci, h in enumerate(headers, 1):
        _cell(ws, 2, ci, h, bold=True, align='center')
    ws.row_dimensions[2].height = 20

    # 데이터
    row = 3
    price_totals = {}  # {price: {'qty': 0, 'amount': 0}}

    for ap in applicants:
        price = ap.get('exercise_price') or 0
        qty   = ap.get('quantity') or 0
        amt   = price * qty

        _cell(ws, row, 1, ap.get('name', ''))
        _cell(ws, row, 2, price,  align='right', number_format='#,##0')
        _cell(ws, row, 3, qty,    align='right', number_format='#,##0')
        _cell(ws, row, 4, amt,    align='right', number_format='#,##0')
        _cell(ws, row, 5, ap.get('broker', '') or '')
        _cell(ws, row, 6, ap.get('account_number', '') or '')
        _cell(ws, row, 7, '')
        ws.row_dimensions[row].height = 18

        if price not in price_totals:
            price_totals[price] = {'qty': 0, 'amount': 0}
        price_totals[price]['qty']    += qty
        price_totals[price]['amount'] += amt

        row += 1

    # 가격별 소계
    row += 1
    ws.merge_cells(f'A{row}:G{row}')
    ws.cell(row=row, column=1).value = '[ 행사가액별 집계 ]'
    ws.cell(row=row, column=1).font = Font(name='맑은 고딕', bold=True, size=11)
    ws.row_dimensions[row].height = 22
    row += 1

    sub_headers = ['행사가액(원)', '수량(주)', '납입금액(원)']
    for ci, h in enumerate(sub_headers, 1):
        _cell(ws, row, ci, h, bold=True, align='center')
    row += 1

    total_qty = 0
    total_amt = 0
    for price in sorted(price_totals.keys()):
        info = price_totals[price]
        _cell(ws, row, 1, price,          align='right', number_format='#,##0')
        _cell(ws, row, 2, info['qty'],    align='right', number_format='#,##0')
        _cell(ws, row, 3, info['amount'], align='right', number_format='#,##0')
        total_qty += info['qty']
        total_amt += info['amount']
        row += 1

    # 합계
    _cell(ws, row, 1, '합계', bold=True, align='center')
    _cell(ws, row, 2, total_qty, bold=True, align='right', number_format='#,##0')
    _cell(ws, row, 3, total_amt, bold=True, align='right', number_format='#,##0')

    # 열 너비
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 12

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


def generate_registration_excel(
        round_name: str,
        exercise_date: str,
        applicants: list,
        reg_config: dict,
        output_path: str) -> str:
    """
    등기신청 집계표 엑셀 생성.
    reg_config keys: reg_date, issue_date, par_value, capital_before,
                     shares_before, company_name, company_reg_num
    applicants: list of dicts with name, exercise_price, quantity, account_number
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '등기신청집계표'

    # ── 스타일 헬퍼 ──────────────────────────────────────────────
    HEADER_FILL  = PatternFill('solid', fgColor='1F4E79')
    SUBHDR_FILL  = PatternFill('solid', fgColor='2E75B6')
    SECTION_FILL = PatternFill('solid', fgColor='D6E4F0')
    WHITE_FONT   = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
    BOLD_FONT    = Font(name='맑은 고딕', bold=True, size=10)
    NORMAL_FONT  = Font(name='맑은 고딕', size=10)
    CENTER       = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT         = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    RIGHT        = Alignment(horizontal='right',  vertical='center')

    def hdr(ws, r, c, val, colspan=1, fill=None, font=None, align=None):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font   = font  or WHITE_FONT
        cell.fill   = fill  or HEADER_FILL
        cell.alignment = align or CENTER
        cell.border = _BORDER
        if colspan > 1:
            ws.merge_cells(start_row=r, start_column=c,
                           end_row=r, end_column=c + colspan - 1)
        return cell

    def dat(ws, r, c, val, bold=False, align=None, fmt=None, colspan=1, fill=None):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font   = BOLD_FONT if bold else NORMAL_FONT
        cell.fill   = fill or PatternFill()
        cell.alignment = align or LEFT
        cell.border = _BORDER
        if fmt:
            cell.number_format = fmt
        if colspan > 1:
            ws.merge_cells(start_row=r, start_column=c,
                           end_row=r, end_column=c + colspan - 1)
        return cell

    # ── 수치 계산 ─────────────────────────────────────────────────
    par_value      = int(reg_config.get('par_value') or 500)
    capital_before = int(reg_config.get('capital_before') or 0)
    shares_before  = int(reg_config.get('shares_before') or 0)

    total_new_shares = sum(ap.get('quantity') or 0 for ap in applicants)
    capital_increase = total_new_shares * par_value
    capital_after    = capital_before + capital_increase
    shares_after     = shares_before  + total_new_shares

    company_name    = reg_config.get('company_name') or 'S2W Inc.'
    company_reg_num = reg_config.get('company_reg_num') or ''
    reg_date        = reg_config.get('reg_date') or ''
    issue_date      = reg_config.get('issue_date') or exercise_date or ''

    # ── 열 너비 ───────────────────────────────────────────────────
    col_widths = [4, 18, 14, 14, 16, 14, 16, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(i)
        ].width = w

    row = 1

    # ── 제목 ─────────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:H{row}')
    tc = ws.cell(row=row, column=1,
                 value=f'스톡옵션 등기신청 집계표  ─  {round_name}')
    tc.font      = Font(name='맑은 고딕', bold=True, size=14, color='1F4E79')
    tc.alignment = CENTER
    ws.row_dimensions[row].height = 32
    row += 1

    # ── 1. 기본 정보 ──────────────────────────────────────────────
    ws.merge_cells(f'A{row}:H{row}')
    sc = ws.cell(row=row, column=1, value='① 기본 정보')
    sc.font      = BOLD_FONT
    sc.fill      = SECTION_FILL
    sc.alignment = LEFT
    sc.border    = _BORDER
    ws.row_dimensions[row].height = 20
    row += 1

    info_rows = [
        ('회사명',        company_name),
        ('법인등록번호',  company_reg_num or '-'),
        ('신주발행일',    issue_date or '-'),
        ('등기신청일',    reg_date or '-'),
        ('액면가 (원)',   f'{par_value:,}'),
    ]
    for label, value in info_rows:
        dat(ws, row, 1, '', colspan=1)
        dat(ws, row, 2, label, bold=True, fill=PatternFill('solid', fgColor='EBF3FB'))
        dat(ws, row, 3, value, colspan=2)
        # 오른쪽 여백
        for c in range(5, 9):
            dat(ws, row, c, '')
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1  # 구분 공백

    # ── 2. 자본금 변동 ─────────────────────────────────────────────
    ws.merge_cells(f'A{row}:H{row}')
    sc2 = ws.cell(row=row, column=1, value='② 자본금 변동 내역')
    sc2.font = BOLD_FONT; sc2.fill = SECTION_FILL
    sc2.alignment = LEFT; sc2.border = _BORDER
    ws.row_dimensions[row].height = 20
    row += 1

    hdr(ws, row, 1, '', fill=SUBHDR_FILL)
    hdr(ws, row, 2, '구분',              fill=SUBHDR_FILL)
    hdr(ws, row, 3, '변경 전',  colspan=2, fill=SUBHDR_FILL)
    hdr(ws, row, 5, '증가',     colspan=2, fill=SUBHDR_FILL)
    hdr(ws, row, 7, '변경 후',  colspan=2, fill=SUBHDR_FILL)
    ws.row_dimensions[row].height = 20
    row += 1

    hdr(ws, row, 1, '', fill=SUBHDR_FILL)
    hdr(ws, row, 2, '', fill=SUBHDR_FILL)
    for c, lbl in [(3,'주식수(주)'),(4,'자본금(원)'),(5,'주식수(주)'),(6,'자본금(원)'),(7,'주식수(주)'),(8,'자본금(원)')]:
        hdr(ws, row, c, lbl, fill=SUBHDR_FILL)
    ws.row_dimensions[row].height = 18
    row += 1

    dat(ws, row, 1, '')
    dat(ws, row, 2, '보통주', bold=True)
    dat(ws, row, 3, shares_before,   align=RIGHT, fmt='#,##0')
    dat(ws, row, 4, capital_before,  align=RIGHT, fmt='#,##0')
    dat(ws, row, 5, total_new_shares,align=RIGHT, fmt='#,##0')
    dat(ws, row, 6, capital_increase,align=RIGHT, fmt='#,##0')
    dat(ws, row, 7, shares_after,    align=RIGHT, fmt='#,##0',
        fill=PatternFill('solid', fgColor='E2EFDA'))
    dat(ws, row, 8, capital_after,   align=RIGHT, fmt='#,##0',
        fill=PatternFill('solid', fgColor='E2EFDA'))
    ws.row_dimensions[row].height = 20
    row += 2

    # ── 3. 신주 발행 내역 ──────────────────────────────────────────
    ws.merge_cells(f'A{row}:H{row}')
    sc3 = ws.cell(row=row, column=1, value='③ 신주 발행 내역 (행사자별)')
    sc3.font = BOLD_FONT; sc3.fill = SECTION_FILL
    sc3.alignment = LEFT; sc3.border = _BORDER
    ws.row_dimensions[row].height = 20
    row += 1

    for c, lbl in enumerate(
            ['No.','성명','행사가액(원)','수량(주)','납입금액(원)','계좌번호','증권사','비고'], 1):
        hdr(ws, row, c, lbl)
    ws.row_dimensions[row].height = 18
    row += 1

    total_qty = 0
    total_amt = 0
    for idx, ap in enumerate(applicants, 1):
        qty   = ap.get('quantity') or 0
        price = ap.get('exercise_price') or 0
        amt   = qty * price
        total_qty += qty
        total_amt += amt
        dat(ws, row, 1, idx,                            align=CENTER)
        dat(ws, row, 2, ap.get('name',''))
        dat(ws, row, 3, price,    align=RIGHT, fmt='#,##0')
        dat(ws, row, 4, qty,      align=RIGHT, fmt='#,##0')
        dat(ws, row, 5, amt,      align=RIGHT, fmt='#,##0')
        dat(ws, row, 6, ap.get('account_number','') or '')
        dat(ws, row, 7, ap.get('broker','') or '')
        dat(ws, row, 8, '')
        ws.row_dimensions[row].height = 18
        row += 1

    # 합계 행
    dat(ws, row, 1, '',      bold=True, fill=PatternFill('solid', fgColor='FFF2CC'))
    dat(ws, row, 2, '합 계', bold=True, fill=PatternFill('solid', fgColor='FFF2CC'))
    dat(ws, row, 3, '',      fill=PatternFill('solid', fgColor='FFF2CC'))
    dat(ws, row, 4, total_qty, bold=True, align=RIGHT, fmt='#,##0',
        fill=PatternFill('solid', fgColor='FFF2CC'))
    dat(ws, row, 5, total_amt, bold=True, align=RIGHT, fmt='#,##0',
        fill=PatternFill('solid', fgColor='FFF2CC'))
    dat(ws, row, 6, '', fill=PatternFill('solid', fgColor='FFF2CC'))
    dat(ws, row, 7, '', fill=PatternFill('solid', fgColor='FFF2CC'))
    dat(ws, row, 8, '', fill=PatternFill('solid', fgColor='FFF2CC'))
    ws.row_dimensions[row].height = 20
    row += 2

    # ── 4. 등기신청 체크리스트 ──────────────────────────────────────
    ws.merge_cells(f'A{row}:H{row}')
    sc4 = ws.cell(row=row, column=1, value='④ 등기신청 구비서류 체크리스트')
    sc4.font = BOLD_FONT; sc4.fill = SECTION_FILL
    sc4.alignment = LEFT; sc4.border = _BORDER
    ws.row_dimensions[row].height = 20
    row += 1

    checklist = [
        ('등기신청서',              '법원 양식, 대표이사 인감 날인'),
        ('주식발행사항 보고서',      '행사내역 전체 기재'),
        ('주주총회/이사회 의사록',   '신주발행 결의 내용 포함'),
        ('납입금 보관증명서',        'Step 03 은행 서류'),
        ('정관',                    '최신 정관 사본'),
        ('법인인감증명서',           '발급 후 3개월 이내'),
        ('대표이사 인감',            '등기신청서에 날인'),
        ('수입인지',                 '등록면허세 영수증 포함'),
    ]
    hdr(ws, row, 1, 'No.',  fill=SUBHDR_FILL)
    hdr(ws, row, 2, '서류명', colspan=3, fill=SUBHDR_FILL)
    hdr(ws, row, 5, '비고',   colspan=3, fill=SUBHDR_FILL)
    hdr(ws, row, 8, '완료',   fill=SUBHDR_FILL)
    ws.row_dimensions[row].height = 18
    row += 1

    for i, (doc, note) in enumerate(checklist, 1):
        dat(ws, row, 1, i,    align=CENTER)
        dat(ws, row, 2, doc,  colspan=3)
        dat(ws, row, 5, note, colspan=3)
        dat(ws, row, 8, '☐',  align=CENTER)
        ws.row_dimensions[row].height = 18
        row += 1

    # ── 시트 설정 ──────────────────────────────────────────────────
    ws.freeze_panes = 'A2'
    ws.sheet_view.showGridLines = True

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


def _load_broker_codes_from_excel():
    """엑셀 파일에서 증권사 코드 로드 (캐싱)."""
    import openpyxl
    excel_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'code.xlsx')

    if not os.path.isfile(excel_path):
        return {}

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        broker_codes = {}

        for i in range(2, ws.max_row + 1):  # Skip header
            row = list(ws[i])
            if len(row) >= 2:
                code = row[0].value
                name = row[1].value
                if code and name:
                    try:
                        code_int = int(float(code))
                        name_str = str(name).strip()
                        broker_codes[name_str] = code_int
                    except:
                        pass

        wb.close()
        return broker_codes
    except Exception as e:
        print(f"[WARNING] 증권사 코드 엑셀 로드 실패: {e}")
        return {}


# 전역 캐시
_BROKER_CODES_CACHE = None


def _get_broker_codes():
    """증권사 코드 가져오기 (엑셀 우선, 실패 시 기본 맵)."""
    global _BROKER_CODES_CACHE

    if _BROKER_CODES_CACHE is None:
        # 엑셀에서 로드 시도
        excel_codes = _load_broker_codes_from_excel()
        if excel_codes:
            _BROKER_CODES_CACHE = excel_codes
            print(f"[INFO] 엑셀에서 {len(excel_codes)}개 증권사 코드 로드 완료")
        else:
            # 기본 맵 사용
            _BROKER_CODES_CACHE = BROKER_CODE_MAP
            print(f"[INFO] 기본 증권사 코드 맵 사용 ({len(_BROKER_CODES_CACHE)}개)")

    return _BROKER_CODES_CACHE


def _broker_to_code(broker_name):
    """증권사 이름을 계좌관리번호(숫자)로 변환 (부분 매칭 지원)."""
    if not broker_name:
        return ''

    broker_codes = _get_broker_codes()
    name = broker_name.strip()

    # 1. 완전 일치
    if name in broker_codes:
        return broker_codes[name]

    # 2. 부분 매칭 (키워드)
    # 예: "미래에셋증권" 입력 → "미래에셋" 키 찾기
    # 예: "미래에셋" 입력 → "미래에셋증권" 키 찾기
    for key, code in broker_codes.items():
        # 입력값이 키에 포함되거나, 키가 입력값에 포함
        if key in name or name in key:
            # 가장 긴 매칭 우선 (예: "NH투자증권" vs "NH")
            return code

    # 3. 더 느슨한 매칭 (공백/특수문자 제거)
    name_clean = name.replace(' ', '').replace('-', '').replace('(주)', '').replace('주식회사', '')
    for key, code in broker_codes.items():
        key_clean = key.replace(' ', '').replace('-', '').replace('(주)', '').replace('주식회사', '')
        if key_clean in name_clean or name_clean in key_clean:
            return code

    return ''


def generate_issuance_detail_excel(
        applicants: list,
        price: int,
        stock_code: str,
        output_path: str) -> str:
    """
    붙임1 일괄발행등록 세부내역 엑셀 생성 (행사가액별).
    예탁원 제출 양식: 구분값 | 발행인관리계좌번호 | ... | 종목코드 | 발행횟수 |
    회차 | 계좌관리번호 | ... | 계좌번호 | 주주명 | 실명번호 | 주식수
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '일괄발행등록세부내역'

    HEADER_FONT = Font(name='맑은 고딕', bold=True, size=9)
    DATA_FONT   = Font(name='맑은 고딕', size=9)
    CENTER      = Alignment(horizontal='center', vertical='center')
    LEFT        = Alignment(horizontal='left', vertical='center')

    headers = ['', '', '종목코드',
               '발행횟수', '회차', '계좌관리번호', '', '',
               '계좌번호', '주주명', '실명번호', '주식수']
    for ci, label in enumerate(headers, 1):
        cell = ws.cell(1, ci, label)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = _BORDER

    # 매칭 실패 로그만 수집 (성공은 로그 안 함)
    failed_matches = []

    for idx, ap in enumerate(applicants):
        r = idx + 2
        broker_name = ap.get('broker', '')
        broker_code = _broker_to_code(broker_name)

        # 실패한 경우만 수집
        if not broker_code and broker_name:
            failed_matches.append(f"{ap.get('name', '?')} (증권사: '{broker_name}')")

        # 계좌번호: OCR 우선, 없으면 수동입력값 사용, 둘 다 없으면 빈칸
        ocr_account = ap.get('ocr_account') or ''
        manual_account = ap.get('account_number') or ''

        # "account" 같은 잘못된 값 필터링
        if ocr_account and ocr_account.lower() in ['account', 'none', 'null']:
            ocr_account = ''
        if manual_account and manual_account.lower() in ['account', 'none', 'null']:
            manual_account = ''

        account = ocr_account or manual_account

        vals = [
            '0001',                                        # A (기존 C)
            50,                                            # B (기존 D)
            int(stock_code),                               # C (기존 E)
            '',                                            # D (기존 F)
            '',                                            # E (기존 G)
            broker_code,                                   # F (기존 H)
            '0000',                                        # G (기존 I)
            '02',                                          # H (기존 J)
            account,                                       # I (기존 K, OCR or 수동입력)
            ap.get('name', ''),                            # J (기존 L)
            ap.get('rrn', '') or '',                       # K (기존 M, OCR 추출)
            ap.get('quantity', 0) or 0,                    # L (기존 N)
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(r, ci, val)
            cell.font = DATA_FONT
            cell.alignment = CENTER if ci != 10 else LEFT  # 주주명(J열)만 왼쪽 정렬
            cell.border = _BORDER
            if ci == 12:  # 주식수 콤마 포맷 (L열)
                cell.number_format = '#,##0'

    # 매칭 실패한 경우 로그 출력
    if failed_matches:
        print(f"    [계좌관리번호 매칭 실패] {len(failed_matches)}명: {', '.join(failed_matches[:3])}" +
              (f" 외 {len(failed_matches)-3}명" if len(failed_matches) > 3 else ""))

    widths = {'A': 6, 'B': 4, 'C': 10, 'D': 8,
              'E': 6, 'F': 14, 'G': 6, 'H': 4, 'I': 18, 'J': 20,
              'K': 18, 'L': 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # 계좌관리번호 참조 시트
    ws2 = wb.create_sheet('계좌관리번호')
    ws2.cell(1, 1, '계좌관리번호 현황').font = Font(name='맑은 고딕', bold=True, size=10)
    seen = set()
    ref_row = 2
    for name, code in sorted(BROKER_CODE_MAP.items(), key=lambda x: x[1]):
        if code not in seen and len(name) > 2:
            ws2.cell(ref_row, 1, code).font = DATA_FONT
            ws2.cell(ref_row, 2, name).font = DATA_FONT
            seen.add(code)
            ref_row += 1
    ws2.column_dimensions['A'].width = 6
    ws2.column_dimensions['B'].width = 20

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path
